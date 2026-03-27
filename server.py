from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any, Set
import uuid
from datetime import datetime, timedelta
import jwt
import bcrypt
from enum import Enum
from bson import ObjectId
import io
import base64
import hashlib
import pyotp
import qrcode
from reportlab.lib import colors as pdf_colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import FCM service
from fcm_service import (
    is_fcm_configured, 
    send_push_notification, 
    send_push_to_multiple,
    send_topic_notification,
    get_fcm_status
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'ambulance_db')]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'ambulancias-sten-secret-key-2024-secure')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

app = FastAPI(title="Ambulancias STEN - Sistema de Gestión")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Enums
class UserRole(str, Enum):
    ADMINISTRADOR = "administrador"
    OPERADOR = "operador"
    COORDINADOR = "coordinador"
    ADMINISTRATIVO = "administrativo"
    PARAMEDICO = "paramedico"
    SUPERVISOR = "supervisor"

# Roles con acceso total
FULL_ACCESS_ROLES = {
    UserRole.ADMINISTRADOR.value,
    UserRole.COORDINADOR.value,
    UserRole.SUPERVISOR.value,
}

# Roles que pueden administrar usuarios
ADMIN_ROLES = {
    UserRole.ADMINISTRADOR.value,
    UserRole.COORDINADOR.value,
    UserRole.SUPERVISOR.value,
}

# Roles que requieren 2FA obligatorio
TWO_FA_REQUIRED_ROLES = {
    UserRole.ADMINISTRADOR.value,
    UserRole.COORDINADOR.value,
    UserRole.SUPERVISOR.value,
}

class ServiceType(str, Enum):
    URGENTE = "urgente"
    PROGRAMADO = "programado"

class ServiceStatus(str, Enum):
    PENDIENTE = "pendiente"
    CONFIRMADO = "confirmado"
    EN_CAMINO = "en_camino"
    EN_CURSO = "en_curso"
    FINALIZADO = "finalizado"
    CANCELADO = "cancelado"

class PaymentType(str, Enum):
    EFECTIVO = "efectivo"
    TRANSFERENCIA = "transferencia"
    PENDIENTE = "pendiente"

class AmbulanceStatus(str, Enum):
    DISPONIBLE = "disponible"
    EN_SERVICIO = "en_servicio"
    MANTENIMIENTO = "mantenimiento"
    FUERA_SERVICIO = "fuera_servicio"

class ExpenseCategory(str, Enum):
    GASOLINA = "gasolina"
    VIATICOS = "viaticos"
    BONOS = "bonos"
    MANTENIMIENTO = "mantenimiento"
    LIMPIEZA = "limpieza"
    FACTURAS = "facturas"
    COMPRAS = "compras"
    OTROS = "otros"

# ====== CONFIGURACIÓN DEL SISTEMA ======
class SystemConfig:
    MAX_ACTIVE_USERS = 20  # Configurable por admin
    PASSWORD_EXPIRY_DAYS = 90
    PASSWORD_EXPIRY_WARNING_DAYS = 7
    PASSWORD_HISTORY_COUNT = 3
    MAX_FAILED_LOGIN_ATTEMPTS = 5
    LOCKOUT_DURATION_MINUTES = 30
    SESSION_TIMEOUT_HOURS = 24
    MAX_SESSIONS_PER_USER = 5
    TWO_FA_APP_NAME = "CareFleet-STEN"
    TWO_FA_ISSUER = "Ambulancias STEN"

# Pydantic Models
class UserBase(BaseModel):
    email: EmailStr
    full_name: str
    role: UserRole = UserRole.OPERADOR
    phone: Optional[str] = None

class UserCreate(UserBase):
    password: str

class UserCreateAdmin(UserBase):
    """Modelo para crear usuarios desde panel de administración"""
    password: str
    assigned_ambulance_id: Optional[str] = None
    is_active: bool = True
    two_fa_required: Optional[bool] = None  # Si None, usa el default del rol

class UserUpdate(BaseModel):
    """Modelo para actualizar usuarios"""
    email: Optional[EmailStr] = None
    full_name: Optional[str] = None
    role: Optional[UserRole] = None
    phone: Optional[str] = None
    assigned_ambulance_id: Optional[str] = None
    is_active: Optional[bool] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str
    two_fa_code: Optional[str] = None

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

class PasswordReset(BaseModel):
    new_password: str

class TwoFASetup(BaseModel):
    code: str  # Código TOTP para verificar

class TwoFADisable(BaseModel):
    password: str
    code: str

class User(UserBase):
    id: str
    created_at: datetime
    is_active: bool = True
    assigned_ambulance_id: Optional[str] = None
    two_fa_enabled: bool = False
    must_change_password: bool = False
    password_expires_at: Optional[datetime] = None
    last_login: Optional[datetime] = None

class UserFull(User):
    """Usuario completo con todos los campos (solo para admin)"""
    failed_login_attempts: int = 0
    locked_until: Optional[datetime] = None
    last_login_ip: Optional[str] = None
    created_by: Optional[str] = None
    updated_at: Optional[datetime] = None
    updated_by: Optional[str] = None
    two_fa_verified: bool = False
    active_sessions_count: int = 0

class UserSession(BaseModel):
    id: str
    user_id: str
    ip_address: Optional[str] = None
    device_info: Optional[str] = None
    created_at: datetime
    last_activity: datetime
    expires_at: datetime
    is_active: bool = True

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: User
    requires_2fa: bool = False
    requires_password_change: bool = False
    password_expiring_soon: bool = False

class PasswordHistory(BaseModel):
    password_hash: str
    created_at: datetime

# Location Models
class Location(BaseModel):
    address: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    captured_at: Optional[datetime] = None

# Ambulance Models
class AmbulanceBase(BaseModel):
    plate_number: str
    unit_number: str
    model: str
    year: int
    status: AmbulanceStatus = AmbulanceStatus.DISPONIBLE
    last_maintenance: Optional[datetime] = None
    next_maintenance: Optional[datetime] = None
    current_km: Optional[int] = None

class AmbulanceCreate(AmbulanceBase):
    pass

class Ambulance(AmbulanceBase):
    id: str
    created_at: datetime
    services_count: int = 0

# Personnel Models
class PersonnelBase(BaseModel):
    full_name: str
    role: str
    phone: str
    is_available: bool = True
    license_number: Optional[str] = None
    license_expiry: Optional[datetime] = None

class PersonnelCreate(PersonnelBase):
    pass

class Personnel(PersonnelBase):
    id: str
    created_at: datetime
    services_count: int = 0

# Patient Models
class PatientInfo(BaseModel):
    name: str
    birth_date: Optional[str] = None  # Fecha de nacimiento
    age: Optional[int] = None  # Calculado automáticamente
    sex: Optional[str] = None  # Sexo
    weight: Optional[float] = None  # Peso en kg
    phone: Optional[str] = None
    diagnosis: Optional[str] = None  # Dx - Diagnóstico
    chronic_diseases: Optional[str] = None  # Enfermedades crónicas
    treating_doctor: Optional[str] = None  # Médico tratante
    room_number: Optional[str] = None  # No. de habitación
    condition: Optional[str] = None  # Legacy field

# Estado del paciente
class PatientStatus(str, Enum):
    ESTABLE = "estable"
    INESTABLE = "inestable"
    CRITICO = "critico"

# Motivo del servicio
class ServiceReason(str, Enum):
    ESTUDIO = "estudio"
    OTRO_HOSPITAL = "otro_hospital"
    DOMICILIO = "domicilio"

# Equipo requerido
class EquipmentRequired(BaseModel):
    oxygen: bool = False
    oxygen_liters: Optional[float] = None  # Obligatorio si oxygen=True
    monitor_oximeter: bool = False
    ventilator: bool = False
    infusion_pumps: bool = False
    infusion_pumps_count: Optional[int] = None  # Obligatorio si infusion_pumps=True

# Personal requerido
class PersonnelRequired(BaseModel):
    doctor: bool = False
    doctor_name: Optional[str] = None  # Obligatorio si doctor=True
    paramedic: bool = False

# Cotización del servicio
class ServiceQuote(BaseModel):
    base_cost: float
    additional_charges: float = 0
    total_estimated: float
    quote_notes: Optional[str] = None

# Ajuste de costo
class CostAdjustment(BaseModel):
    original_amount: float
    new_amount: float
    adjustment_reason: str
    adjusted_by: str
    adjusted_at: datetime

# Service Models - Extended
class ServiceBase(BaseModel):
    service_type: ServiceType
    patient: PatientInfo
    patient_status: Optional[PatientStatus] = None  # Estado del paciente
    service_reason: Optional[ServiceReason] = None  # Motivo del servicio
    study_to_perform: Optional[str] = None  # Estudio a realizar (obligatorio si motivo=estudio)
    origin: str
    origin_location: Optional[Location] = None
    destination: str
    destination_location: Optional[Location] = None
    destination_hospital_name: Optional[str] = None  # Nombre del hospital/lugar
    destination_area: Optional[str] = None  # Área receptora
    scheduled_date: datetime
    request_date: Optional[datetime] = None  # Fecha de solicitud
    request_time: Optional[str] = None  # Hora de solicitud
    equipment_required: Optional[EquipmentRequired] = None
    personnel_required: Optional[PersonnelRequired] = None
    # Datos administrativos
    hospital_account: Optional[str] = None  # Cuenta del hospital
    cash_payment_familiar: Optional[bool] = False  # Pago en efectivo (familiar)
    request_receiver_name: Optional[str] = None  # Nombre de quien recibe la solicitud
    scheduling_nurse: Optional[str] = None  # Enfermera que programa
    # Cotización
    quote: Optional[ServiceQuote] = None
    notes: Optional[str] = None
    requester_name: Optional[str] = None
    requester_phone: Optional[str] = None
    requester_company: Optional[str] = None

class ServiceCreate(ServiceBase):
    pass

class ServiceUpdate(BaseModel):
    status: Optional[ServiceStatus] = None
    ambulance_id: Optional[str] = None
    personnel_ids: Optional[List[str]] = None
    notes: Optional[str] = None
    payment_type: Optional[PaymentType] = None
    payment_amount: Optional[float] = None
    frap_notes: Optional[str] = None
    origin_location: Optional[Location] = None
    destination_location: Optional[Location] = None
    # Transfer payment details
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    account_holder: Optional[str] = None
    transfer_reference: Optional[str] = None
    payment_proof_base64: Optional[str] = None
    # Cost adjustments
    cost_adjustment: Optional[CostAdjustment] = None

class Service(ServiceBase):
    id: str
    status: ServiceStatus = ServiceStatus.PENDIENTE
    ambulance_id: Optional[str] = None
    personnel_ids: List[str] = []
    payment_type: PaymentType = PaymentType.PENDIENTE
    payment_amount: Optional[float] = None
    frap_notes: Optional[str] = None
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    account_holder: Optional[str] = None
    transfer_reference: Optional[str] = None
    payment_proof_base64: Optional[str] = None
    created_at: datetime
    updated_at: datetime
    created_by: Optional[str] = None
    status_history: List[Dict[str, Any]] = []

# Notification Models
class NotificationType(str, Enum):
    SERVICIO_NUEVO = "servicio_nuevo"
    SERVICIO_ASIGNADO = "servicio_asignado"
    SERVICIO_ACTUALIZADO = "servicio_actualizado"
    URGENCIA = "urgencia"
    ALERTA_UNIDADES = "alerta_unidades"
    ALERTA_INVENTARIO = "alerta_inventario"
    ALERTA_CHECKLIST = "alerta_checklist"
    ALERTA_OXIGENO = "alerta_oxigeno"
    UNIDAD_FUERA_SERVICIO = "unidad_fuera_servicio"
    # Nuevos tipos para triggers automáticos
    STOCK_BAJO = "stock_bajo"
    OXIGENO_BAJO = "oxigeno_bajo"
    OXIGENO_CRITICO = "oxigeno_critico"
    ITEM_POR_VENCER = "item_por_vencer"
    AJUSTE_COSTO = "ajuste_costo"
    AJUSTE_INVENTARIO = "ajuste_inventario"
    PACIENTE_CRITICO = "paciente_critico"

class NotificationPriority(str, Enum):
    NORMAL = "normal"
    ALERTA = "alerta"
    CRITICA = "critica"

class Notification(BaseModel):
    id: str
    user_id: str
    title: str
    message: str
    notification_type: NotificationType
    priority: NotificationPriority = NotificationPriority.NORMAL
    service_id: Optional[str] = None
    entity_id: Optional[str] = None  # For linking to ambulance, inventory item, etc.
    entity_type: Optional[str] = None
    is_read: bool = False
    created_at: datetime

# FCM Device Token Model (preparado para Firebase)
class DeviceToken(BaseModel):
    id: str
    user_id: str
    token: str
    device_type: str  # "android", "ios", "web"
    is_active: bool = True
    created_at: datetime
    updated_at: datetime

# Finance Models
class FinanceEntry(BaseModel):
    id: str
    entry_type: str
    amount: float
    category: Optional[str] = None
    description: str
    payment_type: Optional[str] = None
    service_id: Optional[str] = None
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    account_holder: Optional[str] = None
    transfer_reference: Optional[str] = None
    receipt_base64: Optional[str] = None
    created_at: datetime
    created_by: str

class FinanceEntryCreate(BaseModel):
    entry_type: str
    amount: float
    category: Optional[str] = None
    description: str
    payment_type: Optional[str] = None
    service_id: Optional[str] = None
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    account_holder: Optional[str] = None
    transfer_reference: Optional[str] = None
    receipt_base64: Optional[str] = None

# Checklist Models
class ChecklistItemExtended(BaseModel):
    name: str
    status: bool  # True = OK, False = Falla
    observation: Optional[str] = None
    critical: Optional[bool] = False

class ChecklistItem(BaseModel):
    name: str
    status: bool  # True = OK, False = Falla
    notes: Optional[str] = None

class ChecklistLevel(BaseModel):
    name: str
    level: str  # "bueno", "malo", "bajo", "normal", "alto"
    observation: Optional[str] = None
    critical: Optional[bool] = False

class ChecklistCreate(BaseModel):
    ambulance_id: str
    operator_id: str
    operator_name: Optional[str] = None
    paramedic_id: Optional[str] = None
    shift: str  # "matutino", "vespertino", "nocturno"
    date: datetime
    km: int
    fuel_level: str  # "vacio", "1/4", "1/2", "3/4", "lleno"
    # Sections - supporting both old and new format
    apariencia_general: List[dict]
    cabina_operadores: List[dict]
    compartimiento_motor: List[dict]
    niveles: List[dict]
    exterior_operador: List[dict]
    zona_frontal: List[dict]
    exterior_copiloto: List[dict]
    compartimento_paciente: List[dict]
    zona_posterior: List[dict]
    herramientas: List[dict]
    saldo_gasolina: float
    observations: Optional[str] = None
    photos_base64: List[str] = []
    signature_base64: Optional[str] = None
    completed_at: Optional[datetime] = None
    failure_count: Optional[int] = 0

class Checklist(ChecklistCreate):
    id: str
    created_at: datetime
    has_failures: bool = False
    failure_count: int = 0

# Inventory Models - Sistema de Inventario por Ambulancia Tipo Checklist

# Categorías de inventario
class InventoryCategory(str, Enum):
    BOTIQUIN_TRAUMA = "botiquin_trauma"
    BOTIQUIN_VIA_AEREA = "botiquin_via_aerea"
    OXIGENO = "oxigeno"
    MEDICAMENTOS = "medicamentos"
    RACK = "rack"
    CABINA_ATENCION = "cabina_atencion"
    EQUIPO_PORTATIL = "equipo_portatil"
    MONITOR_SIGNOS_VITALES = "monitor_signos_vitales"
    OTROS = "otros"

# Tipos de tanque de oxígeno
class OxygenTankType(str, Enum):
    M = "M"  # Estacionario UM05 - 3500L
    K = "K"  # Estacionario UM03 - 9500L
    D = "D"  # Portátil - 680L
    C = "C"  # Portátil - 680L

# Estado del tanque de oxígeno (semáforo)
class OxygenTankStatus(str, Enum):
    VERDE = "verde"      # > 800 litros
    AMARILLO = "amarillo"  # 300-800 litros
    ROJO = "rojo"        # <= 200 litros
    CRITICO = "critico"  # 0 litros

# Modelo de tanque de oxígeno
class OxygenTank(BaseModel):
    id: str
    ambulance_id: str
    tank_type: OxygenTankType
    is_portable: bool
    capacity_liters: int  # Capacidad total en litros
    max_psi: int = 2000
    current_psi: int
    current_liters: int  # Calculado automáticamente
    status: OxygenTankStatus
    last_refill_date: Optional[datetime] = None
    last_movement_by: Optional[str] = None
    created_at: datetime
    updated_at: datetime

class OxygenTankCreate(BaseModel):
    ambulance_id: str
    tank_type: OxygenTankType
    is_portable: bool
    capacity_liters: int
    current_psi: int

# Movimiento de oxígeno
class OxygenMovementType(str, Enum):
    UTILIZADO = "utilizado"
    RECARGADO = "recargado"
    REEMPLAZADO = "reemplazado"

class OxygenMovement(BaseModel):
    id: str
    tank_id: str
    ambulance_id: str
    movement_type: OxygenMovementType
    psi_before: int
    psi_after: int
    liters_used: Optional[int] = None
    liters_added: Optional[int] = None
    observations: Optional[str] = None
    user_id: str
    user_name: str
    created_at: datetime

# Item de inventario estructurado
class InventoryItemBase(BaseModel):
    name: str
    category: InventoryCategory
    expected_quantity: int  # Cantidad esperada (configurable)
    current_quantity: int
    unit: str = "piezas"
    expiry_date: Optional[datetime] = None
    lot_number: Optional[str] = None
    observations: Optional[str] = None

class InventoryItemCreate(InventoryItemBase):
    ambulance_id: str

class InventoryItem(InventoryItemBase):
    id: str
    ambulance_id: str
    has_difference: bool = False  # Si actual != esperado
    is_below_minimum: bool = False
    is_expiring_soon: bool = False  # 30 días antes
    created_at: datetime
    updated_at: datetime

# Movimiento de inventario
class InventoryMovementType(str, Enum):
    ENTRADA = "entrada"
    SALIDA = "salida"
    AJUSTE = "ajuste"
    CORRECCION = "correccion"

class InventoryMovement(BaseModel):
    id: str
    item_id: str
    item_name: str
    ambulance_id: str
    movement_type: InventoryMovementType
    quantity_before: int
    quantity_moved: int
    quantity_after: int
    reason: str
    observations: Optional[str] = None
    service_id: Optional[str] = None  # Si fue por un servicio
    user_id: str
    user_name: str
    created_at: datetime
    # Los movimientos NO pueden eliminarse, solo agregarse correcciones
    is_correction: bool = False
    correction_justification: Optional[str] = None

# Registro de inventario completo por ambulancia (tipo checklist)
class AmbulanceInventoryRecord(BaseModel):
    id: str
    ambulance_id: str
    ambulance_unit: str
    date: datetime
    shift: str  # "matutino", "vespertino", "nocturno"
    paramedic_id: str
    paramedic_name: str
    oxygen_portable_status: str  # Resumen del estado
    oxygen_stationary_status: str
    items: List[InventoryItem]
    oxygen_tanks: List[OxygenTank]
    total_items: int
    items_with_difference: int
    items_below_minimum: int
    items_expiring_soon: int
    observations: Optional[str] = None
    created_at: datetime

# Audit Log
class AuditLog(BaseModel):
    id: str
    user_id: str
    user_name: str
    action: str
    entity_type: str
    entity_id: str
    details: Dict[str, Any]
    created_at: datetime

# Helper Functions
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    expiration = datetime.utcnow() + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        'user_id': user_id,
        'email': email,
        'role': role,
        'exp': expiration
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get('user_id')
        if not user_id:
            raise HTTPException(status_code=401, detail="Token inválido")
        user = await db.users.find_one({"id": user_id})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        return User(**user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

async def create_notification(
    user_id: str, 
    title: str, 
    message: str, 
    notification_type: NotificationType, 
    priority: NotificationPriority = NotificationPriority.NORMAL,
    service_id: str = None,
    entity_id: str = None,
    entity_type: str = None
):
    notification = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "title": title,
        "message": message,
        "notification_type": notification_type,
        "priority": priority,
        "service_id": service_id,
        "entity_id": entity_id,
        "entity_type": entity_type,
        "is_read": False,
        "created_at": datetime.utcnow()
    }
    await db.notifications.insert_one(notification)
    
# Enviar push notification via FCM
    try:
        fcm_priority = "normal"
        if priority in [NotificationPriority.ALERTA, NotificationPriority.CRITICA]:
            fcm_priority = "critica" if priority == NotificationPriority.CRITICA else "alerta"
        
        device_tokens = await db.device_tokens.find({
            "user_id": user_id,
            "is_active": True
        }).to_list(10)
        
        tokens = [dt["token"] for dt in device_tokens]
        
        if tokens:
            await send_push_to_multiple(
                tokens=tokens,
                title=title,
                body=message,
                priority=fcm_priority,
                data={
                    "notification_type": str(notification_type),
                    "service_id": service_id or "",
                    "entity_id": entity_id or "",
                }
            )
    except Exception as e:
        print(f"[FCM] Error enviando push: {e}")
    
    return notification
async def notify_all_coordinators(
    title: str, 
    message: str, 
    notification_type: NotificationType, 
    priority: NotificationPriority = NotificationPriority.NORMAL,
    service_id: str = None,
    entity_id: str = None,
    entity_type: str = None
):
    coordinators = await db.users.find({"role": {"$in": ["coordinador", "supervisor"]}}).to_list(100)
    for coord in coordinators:
        await create_notification(
            coord["id"], title, message, notification_type, 
            priority, service_id, entity_id, entity_type
        )

async def notify_by_role(
    roles: List[str],
    title: str, 
    message: str, 
    notification_type: NotificationType, 
    priority: NotificationPriority = NotificationPriority.NORMAL,
    service_id: str = None,
    entity_id: str = None,
    entity_type: str = None
):
    """Notificar a usuarios por rol específico"""
    users = await db.users.find({"role": {"$in": roles}}).to_list(100)
    for user in users:
        await create_notification(
            user["id"], title, message, notification_type, 
            priority, service_id, entity_id, entity_type
        )

async def log_audit(user_id: str, user_name: str, action: str, entity_type: str, entity_id: str, details: dict = None):
    audit = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "user_name": user_name,
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "details": details or {},
        "created_at": datetime.utcnow()
    }
    await db.audit_logs.insert_one(audit)

async def check_available_ambulances():
    """Check if there are available ambulances and create alert if none"""
    available = await db.ambulances.count_documents({"status": "disponible"})
    if available == 0:
        await notify_all_coordinators(
            "⚠️ ALERTA: Sin unidades disponibles",
            "No hay ambulancias disponibles en este momento. Todas las unidades están en servicio o mantenimiento.",
            NotificationType.ALERTA_UNIDADES
        )
    return available

# ====== TRIGGERS DE NOTIFICACIONES AUTOMÁTICAS ======

async def check_low_stock_items():
    """Verificar items con stock bajo y generar alertas"""
    # Buscar items donde la cantidad actual es menor al 20% de la esperada
    items = await db.inventory.find({
        "$expr": {"$lt": ["$current_quantity", {"$multiply": ["$expected_quantity", 0.2]}]}
    }).to_list(100)
    
    for item in items:
        ambulance = await db.ambulances.find_one({"id": item["ambulance_id"]})
        unit = ambulance.get("unit_number", "N/A") if ambulance else "N/A"
        
        await notify_all_coordinators(
            f"📦 Stock bajo - {unit}",
            f"El item '{item['name']}' tiene solo {item['current_quantity']}/{item['expected_quantity']} unidades",
            NotificationType.STOCK_BAJO,
            NotificationPriority.ALERTA,
            entity_id=item["id"],
            entity_type="inventario"
        )
    
    return len(items)

async def check_oxygen_levels():
    """Verificar niveles de oxígeno y generar alertas según semáforo"""
    tanks = await db.oxygen_tanks.find().to_list(100)
    alerts_generated = 0
    
    for tank in tanks:
        current_liters = tank.get("current_liters", 0)
        ambulance = await db.ambulances.find_one({"id": tank["ambulance_id"]})
        unit = ambulance.get("unit_number", "N/A") if ambulance else "N/A"
        tank_type = tank.get("tank_type", "N/A")
        
        # Umbral crítico: 200 litros
        # Alerta: 300-800 litros
        if current_liters <= 0:
            # VACÍO - Crítico inmediato
            await notify_all_coordinators(
                f"🚨 CRÍTICO: Tanque vacío - {unit}",
                f"El tanque {tank_type} está VACÍO. Se requiere recarga inmediata.",
                NotificationType.OXIGENO_CRITICO,
                NotificationPriority.CRITICA,
                entity_id=tank["id"],
                entity_type="oxigeno"
            )
            alerts_generated += 1
        elif current_liters <= 200:
            # Crítico - menos de 200L
            await notify_all_coordinators(
                f"🔴 Oxígeno crítico - {unit}",
                f"El tanque {tank_type} tiene {current_liters}L restantes. Umbral crítico alcanzado.",
                NotificationType.OXIGENO_CRITICO,
                NotificationPriority.CRITICA,
                entity_id=tank["id"],
                entity_type="oxigeno"
            )
            alerts_generated += 1
        elif current_liters <= 800:
            # Alerta - entre 200 y 800L
            await notify_all_coordinators(
                f"🟡 Oxígeno bajo - {unit}",
                f"El tanque {tank_type} tiene {current_liters}L restantes. Programar recarga.",
                NotificationType.OXIGENO_BAJO,
                NotificationPriority.ALERTA,
                entity_id=tank["id"],
                entity_type="oxigeno"
            )
            alerts_generated += 1
    
    return alerts_generated

async def check_expiring_items(days_before: int = 30):
    """Verificar items que vencen en los próximos N días"""
    from_date = datetime.utcnow()
    to_date = from_date + timedelta(days=days_before)
    
    items = await db.inventory.find({
        "expiry_date": {"$gte": from_date, "$lte": to_date}
    }).to_list(100)
    
    for item in items:
        ambulance = await db.ambulances.find_one({"id": item["ambulance_id"]})
        unit = ambulance.get("unit_number", "N/A") if ambulance else "N/A"
        expiry = item.get("expiry_date")
        days_left = (expiry - from_date).days if expiry else 0
        
        priority = NotificationPriority.CRITICA if days_left <= 7 else NotificationPriority.ALERTA
        emoji = "🚨" if days_left <= 7 else "⚠️"
        
        await notify_all_coordinators(
            f"{emoji} Item por vencer - {unit}",
            f"'{item['name']}' vence en {days_left} días. Lote: {item.get('lot_number', 'N/A')}",
            NotificationType.ITEM_POR_VENCER,
            priority,
            entity_id=item["id"],
            entity_type="inventario"
        )
    
    return len(items)

async def trigger_cost_modification_notification(
    service_id: str,
    original_amount: float,
    new_amount: float,
    reason: str,
    modified_by: str
):
    """Notificar cuando se modifica el costo de un servicio"""
    service = await db.services.find_one({"id": service_id})
    if not service:
        return
    
    patient_name = service.get("patient", {}).get("name", "N/A")
    difference = new_amount - original_amount
    sign = "+" if difference > 0 else ""
    
    await notify_all_coordinators(
        f"💰 Ajuste de costo - Servicio",
        f"Paciente: {patient_name}\nOriginal: ${original_amount:.2f} → Nuevo: ${new_amount:.2f} ({sign}${difference:.2f})\nMotivo: {reason}",
        NotificationType.AJUSTE_COSTO,
        NotificationPriority.NORMAL,
        service_id=service_id
    )
    
    # Registrar en auditoría
    await log_audit(
        modified_by, "Sistema", "ajuste_costo", "servicio", service_id,
        {"original": original_amount, "new": new_amount, "reason": reason}
    )

async def trigger_critical_inventory_adjustment(
    item_id: str,
    item_name: str,
    ambulance_id: str,
    adjustment_type: str,
    quantity_change: int,
    reason: str,
    user_name: str
):
    """Notificar cuando hay un ajuste crítico de inventario (cambios grandes)"""
    ambulance = await db.ambulances.find_one({"id": ambulance_id})
    unit = ambulance.get("unit_number", "N/A") if ambulance else "N/A"
    
    # Consideramos crítico si el cambio es mayor a 5 unidades o es una corrección
    is_critical = abs(quantity_change) > 5 or adjustment_type == "correccion"
    
    if is_critical:
        priority = NotificationPriority.ALERTA
        emoji = "⚠️"
    else:
        return  # No notificar ajustes pequeños
    
    await notify_all_coordinators(
        f"{emoji} Ajuste de inventario - {unit}",
        f"Item: {item_name}\nTipo: {adjustment_type}\nCambio: {quantity_change:+d} unidades\nRazón: {reason}\nPor: {user_name}",
        NotificationType.AJUSTE_INVENTARIO,
        priority,
        entity_id=item_id,
        entity_type="inventario"
    )

# Endpoint para ejecutar verificaciones periódicas (puede ser llamado por un cron job)
@api_router.post("/notifications/check-alerts")
async def run_notification_checks(current_user: User = Depends(get_current_user)):
    """Ejecutar todas las verificaciones de alertas automáticas"""
    if current_user.role not in ["coordinador", "supervisor", "administrativo"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    results = {
        "low_stock_alerts": await check_low_stock_items(),
        "oxygen_alerts": await check_oxygen_levels(),
        "expiring_items_alerts": await check_expiring_items(30),
        "checked_at": datetime.utcnow().isoformat()
    }
    
    return results

# Endpoint para configurar días de alerta de vencimiento
@api_router.post("/notifications/check-expiring/{days}")
async def check_expiring_by_days(days: int, current_user: User = Depends(get_current_user)):
    """Verificar items que vencen en N días"""
    if days < 1 or days > 365:
        raise HTTPException(status_code=400, detail="Días debe estar entre 1 y 365")
    
    alerts = await check_expiring_items(days)
    return {"days_checked": days, "alerts_generated": alerts}

# ====== HELPER FUNCTIONS FOR AUTH ======

def hash_token(token: str) -> str:
    """Hash de token para almacenamiento seguro"""
    return hashlib.sha256(token.encode()).hexdigest()

def parse_device_info(user_agent: str) -> str:
    """Extrae información básica del dispositivo del user agent"""
    ua = user_agent.lower() if user_agent else ""
    
    if "windows" in ua:
        os_info = "Windows"
    elif "macintosh" in ua or "mac os" in ua:
        os_info = "macOS"
    elif "linux" in ua:
        os_info = "Linux"
    elif "android" in ua:
        os_info = "Android"
    elif "iphone" in ua or "ipad" in ua:
        os_info = "iOS"
    else:
        os_info = "Desconocido"
    
    if "chrome" in ua and "edg" not in ua:
        browser = "Chrome"
    elif "firefox" in ua:
        browser = "Firefox"
    elif "safari" in ua and "chrome" not in ua:
        browser = "Safari"
    elif "edg" in ua:
        browser = "Edge"
    else:
        browser = "Otro"
    
    return f"{browser} en {os_info}"

def calculate_password_expiry() -> datetime:
    """Calcula fecha de expiración de contraseña"""
    return datetime.utcnow() + timedelta(days=SystemConfig.PASSWORD_EXPIRY_DAYS)

def is_password_expiring_soon(expires_at: Optional[datetime]) -> bool:
    """Verifica si la contraseña expira pronto"""
    if not expires_at:
        return False
    warning_date = datetime.utcnow() + timedelta(days=SystemConfig.PASSWORD_EXPIRY_WARNING_DAYS)
    return expires_at <= warning_date

def is_password_expired(expires_at: Optional[datetime]) -> bool:
    """Verifica si la contraseña ha expirado"""
    if not expires_at:
        return False
    return datetime.utcnow() >= expires_at

def validate_password_strength(password: str) -> tuple:
    """Valida la fortaleza de una contraseña"""
    if len(password) < 8:
        return False, "La contraseña debe tener al menos 8 caracteres"
    if not any(c.isupper() for c in password):
        return False, "La contraseña debe contener al menos una mayúscula"
    if not any(c.islower() for c in password):
        return False, "La contraseña debe contener al menos una minúscula"
    if not any(c.isdigit() for c in password):
        return False, "La contraseña debe contener al menos un número"
    special_chars = "!@#$%^&*()_+-=[]{}|;:,.<>?"
    if not any(c in special_chars for c in password):
        return False, "La contraseña debe contener al menos un carácter especial (!@#$%^&*...)"
    return True, "OK"

def is_password_in_history(password: str, history: list) -> bool:
    """Verifica si la contraseña está en el historial reciente"""
    for entry in history[-SystemConfig.PASSWORD_HISTORY_COUNT:]:
        if verify_password(password, entry.get("password_hash", "")):
            return True
    return False

async def create_session(user_id: str, token: str, request: Request = None):
    """Crea una nueva sesión para el usuario"""
    now = datetime.utcnow()
    
    ip_address = None
    user_agent = None
    
    if request:
        ip_address = request.client.host if request.client else None
        user_agent = request.headers.get("user-agent", "")[:500]
    
    session = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "token_hash": hash_token(token),
        "ip_address": ip_address,
        "user_agent": user_agent,
        "device_info": parse_device_info(user_agent) if user_agent else None,
        "created_at": now,
        "last_activity": now,
        "expires_at": now + timedelta(hours=SystemConfig.SESSION_TIMEOUT_HOURS),
        "is_active": True,
    }
    
    await db.sessions.insert_one(session)
    return session

async def close_session(session_id: str):
    """Cierra una sesión específica"""
    await db.sessions.update_one(
        {"id": session_id},
        {"$set": {"is_active": False, "closed_at": datetime.utcnow()}}
    )

async def close_all_user_sessions(user_id: str, except_session_id: str = None):
    """Cierra todas las sesiones de un usuario"""
    query = {"user_id": user_id, "is_active": True}
    if except_session_id:
        query["id"] = {"$ne": except_session_id}
    
    await db.sessions.update_many(
        query,
        {"$set": {"is_active": False, "closed_at": datetime.utcnow()}}
    )

async def get_user_active_sessions(user_id: str) -> list:
    """Obtiene las sesiones activas de un usuario"""
    now = datetime.utcnow()
    sessions = await db.sessions.find({
        "user_id": user_id,
        "is_active": True,
        "expires_at": {"$gt": now}
    }).to_list(100)
    
    return sessions

async def count_user_active_sessions(user_id: str) -> int:
    """Cuenta las sesiones activas de un usuario"""
    now = datetime.utcnow()
    return await db.sessions.count_documents({
        "user_id": user_id,
        "is_active": True,
        "expires_at": {"$gt": now}
    })

async def log_login_attempt(email: str, success: bool, user_id: str = None, 
                           failure_reason: str = None, request: Request = None):
    """Registra un intento de login"""
    attempt = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "email": email,
        "ip_address": request.client.host if request and request.client else None,
        "user_agent": request.headers.get("user-agent", "")[:500] if request else None,
        "success": success,
        "failure_reason": failure_reason,
        "timestamp": datetime.utcnow(),
    }
    await db.login_attempts.insert_one(attempt)

async def get_system_config():
    """Obtiene la configuración del sistema desde la DB o usa defaults"""
    config = await db.system_config.find_one({"id": "main"})
    if not config:
        config = {
            "id": "main",
            "max_active_users": SystemConfig.MAX_ACTIVE_USERS,
            "password_expiry_days": SystemConfig.PASSWORD_EXPIRY_DAYS,
            "two_fa_optional_for_paramedics": True,
            "created_at": datetime.utcnow()
        }
        await db.system_config.insert_one(config)
    return config

async def count_active_users() -> int:
    """Cuenta usuarios activos"""
    return await db.users.count_documents({"is_active": True})

# Auth Routes
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate, request: Request):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    # Validar contraseña
    valid, msg = validate_password_strength(user_data.password)
    if not valid:
        raise HTTPException(status_code=400, detail=msg)
    
    # Verificar límite de usuarios
    config = await get_system_config()
    active_count = await count_active_users()
    if active_count >= config.get("max_active_users", SystemConfig.MAX_ACTIVE_USERS):
        raise HTTPException(status_code=400, detail="Se ha alcanzado el límite de usuarios activos")
    
    user_id = str(uuid.uuid4())
    now = datetime.utcnow()
    
    # Determinar si requiere 2FA
    requires_2fa = user_data.role.value in TWO_FA_REQUIRED_ROLES
    
    user_dict = {
        "id": user_id,
        "email": user_data.email,
        "full_name": user_data.full_name,
        "role": user_data.role.value,
        "phone": user_data.phone,
        "hashed_password": hash_password(user_data.password),
        "password_history": [{"password_hash": hash_password(user_data.password), "created_at": now}],
        "password_expires_at": calculate_password_expiry(),
        "must_change_password": False,
        "created_at": now,
        "is_active": True,
        "two_fa_enabled": False,
        "two_fa_secret": None,
        "two_fa_verified": False,
        "two_fa_required": requires_2fa,
        "failed_login_attempts": 0,
        "locked_until": None,
        "last_login": None,
        "last_login_ip": None,
        "assigned_ambulance_id": None,
    }
    await db.users.insert_one(user_dict)
    
    token = create_token(user_id, user_data.email, user_data.role.value)
    await create_session(user_id, token, request)
    
    user_response = User(
        id=user_id,
        email=user_data.email,
        full_name=user_data.full_name,
        role=user_data.role,
        phone=user_data.phone,
        created_at=now,
        is_active=True,
        two_fa_enabled=False,
        must_change_password=False,
        password_expires_at=calculate_password_expiry(),
    )
    
    return TokenResponse(
        access_token=token, 
        user=user_response,
        requires_2fa=requires_2fa and not user_dict["two_fa_enabled"],
        requires_password_change=False,
        password_expiring_soon=False
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin, request: Request):
    user = await db.users.find_one({"email": credentials.email})
    
    if not user:
        await log_login_attempt(credentials.email, False, None, "Usuario no encontrado", request)
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    # Verificar si está bloqueado
    if user.get("locked_until"):
        locked_until = user["locked_until"]
        if isinstance(locked_until, str):
            locked_until = datetime.fromisoformat(locked_until.replace("Z", "+00:00"))
        if datetime.utcnow() < locked_until:
            await log_login_attempt(credentials.email, False, user["id"], "Cuenta bloqueada", request)
            raise HTTPException(status_code=401, detail="Cuenta bloqueada temporalmente. Intente más tarde.")
    
    # Verificar si está activo
    if not user.get("is_active", True):
        await log_login_attempt(credentials.email, False, user["id"], "Cuenta desactivada", request)
        raise HTTPException(status_code=401, detail="Cuenta desactivada. Contacte al administrador.")
    
    # Verificar contraseña
    if not verify_password(credentials.password, user["hashed_password"]):
        # Incrementar intentos fallidos
        failed_attempts = user.get("failed_login_attempts", 0) + 1
        update_data = {"failed_login_attempts": failed_attempts}
        
        if failed_attempts >= SystemConfig.MAX_FAILED_LOGIN_ATTEMPTS:
            update_data["locked_until"] = datetime.utcnow() + timedelta(minutes=SystemConfig.LOCKOUT_DURATION_MINUTES)
            await log_login_attempt(credentials.email, False, user["id"], f"Cuenta bloqueada tras {failed_attempts} intentos", request)
        else:
            await log_login_attempt(credentials.email, False, user["id"], "Contraseña incorrecta", request)
        
        await db.users.update_one({"id": user["id"]}, {"$set": update_data})
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    # Verificar 2FA si está habilitado
    if user.get("two_fa_enabled") and user.get("two_fa_verified"):
        if not credentials.two_fa_code:
            return TokenResponse(
                access_token="",
                user=User(**user),
                requires_2fa=True,
                requires_password_change=False,
                password_expiring_soon=False
            )
        
        # Verificar código TOTP
        totp = pyotp.TOTP(user.get("two_fa_secret", ""))
        if not totp.verify(credentials.two_fa_code, valid_window=1):
            await log_login_attempt(credentials.email, False, user["id"], "Código 2FA inválido", request)
            raise HTTPException(status_code=401, detail="Código de autenticación inválido")
    
    # Login exitoso - resetear intentos fallidos
    now = datetime.utcnow()
    ip_address = request.client.host if request.client else None
    
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {
            "failed_login_attempts": 0,
            "locked_until": None,
            "last_login": now,
            "last_login_ip": ip_address
        }}
    )
    
    await log_login_attempt(credentials.email, True, user["id"], None, request)
    
    # Crear token y sesión
    token = create_token(user["id"], user["email"], user["role"])
    await create_session(user["id"], token, request)
    
    # Verificar expiración de contraseña
    password_expires_at = user.get("password_expires_at")
    password_expired = is_password_expired(password_expires_at)
    password_expiring_soon = is_password_expiring_soon(password_expires_at)
    must_change = user.get("must_change_password", False) or password_expired
    
    # Verificar si requiere configurar 2FA
    requires_2fa_setup = (
        user["role"] in TWO_FA_REQUIRED_ROLES and 
        not user.get("two_fa_enabled") and 
        not user.get("two_fa_verified")
    )
    
    user_response = User(
        id=user["id"],
        email=user["email"],
        full_name=user["full_name"],
        role=UserRole(user["role"]),
        phone=user.get("phone"),
        created_at=user["created_at"],
        is_active=user.get("is_active", True),
        assigned_ambulance_id=user.get("assigned_ambulance_id"),
        two_fa_enabled=user.get("two_fa_enabled", False),
        must_change_password=must_change,
        password_expires_at=password_expires_at,
        last_login=now
    )
    
    return TokenResponse(
        access_token=token, 
        user=user_response,
        requires_2fa=requires_2fa_setup,
        requires_password_change=must_change,
        password_expiring_soon=password_expiring_soon
    )

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

@api_router.post("/auth/logout")
async def logout(current_user: User = Depends(get_current_user), request: Request = None):
    """Cierra la sesión actual"""
    # Cerrar todas las sesiones del usuario con el token actual
    if request:
        auth_header = request.headers.get("authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
            token_hash = hash_token(token)
            await db.sessions.update_one(
                {"user_id": current_user.id, "token_hash": token_hash},
                {"$set": {"is_active": False, "closed_at": datetime.utcnow()}}
            )
    
    await log_audit(current_user.id, current_user.full_name, "logout", "session", None)
    return {"message": "Sesión cerrada correctamente"}

# ====== 2FA ENDPOINTS ======

@api_router.post("/auth/2fa/setup")
async def setup_2fa(current_user: User = Depends(get_current_user)):
    """Inicia el proceso de configuración de 2FA"""
    # Generar secreto TOTP
    secret = pyotp.random_base32()
    
    # Generar URI para app autenticadora
    totp = pyotp.TOTP(secret)
    uri = totp.provisioning_uri(
        name=current_user.email,
        issuer_name=SystemConfig.TWO_FA_ISSUER
    )
    
    # Generar QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(uri)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    qr_base64 = f"data:image/png;base64,{base64.b64encode(buffer.getvalue()).decode()}"
    
    # Guardar secreto temporalmente
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"two_fa_secret": secret, "two_fa_enabled": False, "two_fa_verified": False}}
    )
    
    return {
        "secret": secret,
        "qr_code": qr_base64,
        "uri": uri,
        "message": "Escanea el código QR con tu app autenticadora y verifica con un código"
    }

@api_router.post("/auth/2fa/verify")
async def verify_2fa(data: TwoFASetup, current_user: User = Depends(get_current_user)):
    """Verifica y activa 2FA"""
    user = await db.users.find_one({"id": current_user.id})
    if not user or not user.get("two_fa_secret"):
        raise HTTPException(status_code=400, detail="Primero debe configurar 2FA")
    
    totp = pyotp.TOTP(user["two_fa_secret"])
    if not totp.verify(data.code, valid_window=1):
        raise HTTPException(status_code=400, detail="Código inválido")
    
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"two_fa_enabled": True, "two_fa_verified": True}}
    )
    
    await log_audit(current_user.id, current_user.full_name, "two_fa_enabled", "user", current_user.id)
    
    return {"message": "2FA activado correctamente"}

@api_router.post("/auth/2fa/disable")
async def disable_2fa(data: TwoFADisable, current_user: User = Depends(get_current_user)):
    """Desactiva 2FA"""
    user = await db.users.find_one({"id": current_user.id})
    
    # Verificar si el rol requiere 2FA obligatorio
    if current_user.role.value in TWO_FA_REQUIRED_ROLES:
        raise HTTPException(status_code=403, detail="2FA es obligatorio para tu rol")
    
    # Verificar contraseña
    if not verify_password(data.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Contraseña incorrecta")
    
    # Verificar código TOTP
    if user.get("two_fa_enabled"):
        totp = pyotp.TOTP(user.get("two_fa_secret", ""))
        if not totp.verify(data.code, valid_window=1):
            raise HTTPException(status_code=400, detail="Código 2FA inválido")
    
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"two_fa_enabled": False, "two_fa_verified": False, "two_fa_secret": None}}
    )
    
    await log_audit(current_user.id, current_user.full_name, "two_fa_disabled", "user", current_user.id)
    
    return {"message": "2FA desactivado"}

# ====== PASSWORD MANAGEMENT ======

@api_router.post("/auth/password/change")
async def change_password(data: PasswordChange, current_user: User = Depends(get_current_user)):
    """Cambia la contraseña del usuario actual"""
    user = await db.users.find_one({"id": current_user.id})
    
    # Verificar contraseña actual
    if not verify_password(data.current_password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Contraseña actual incorrecta")
    
    # Validar nueva contraseña
    valid, msg = validate_password_strength(data.new_password)
    if not valid:
        raise HTTPException(status_code=400, detail=msg)
    
    # Verificar que no esté en el historial
    password_history = user.get("password_history", [])
    if is_password_in_history(data.new_password, password_history):
        raise HTTPException(status_code=400, detail="No puede reutilizar las últimas 3 contraseñas")
    
    # Actualizar contraseña
    new_hash = hash_password(data.new_password)
    password_history.append({"password_hash": new_hash, "created_at": datetime.utcnow()})
    
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {
            "hashed_password": new_hash,
            "password_history": password_history[-SystemConfig.PASSWORD_HISTORY_COUNT:],
            "password_expires_at": calculate_password_expiry(),
            "must_change_password": False
        }}
    )
    
    await log_audit(current_user.id, current_user.full_name, "password_changed", "user", current_user.id)
    
    return {"message": "Contraseña actualizada correctamente"}

# ====== SESSION MANAGEMENT ======

@api_router.get("/auth/sessions")
async def get_my_sessions(current_user: User = Depends(get_current_user)):
    """Obtiene las sesiones activas del usuario actual"""
    sessions = await get_user_active_sessions(current_user.id)
    return [{
        "id": s["id"],
        "ip_address": s.get("ip_address"),
        "device_info": s.get("device_info"),
        "created_at": s["created_at"],
        "last_activity": s.get("last_activity"),
        "is_current": False  # Se puede mejorar comparando con el token actual
    } for s in sessions]

@api_router.delete("/auth/sessions/{session_id}")
async def close_session_endpoint(session_id: str, current_user: User = Depends(get_current_user)):
    """Cierra una sesión específica"""
    session = await db.sessions.find_one({"id": session_id, "user_id": current_user.id})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    
    await close_session(session_id)
    await log_audit(current_user.id, current_user.full_name, "session_closed_remote", "session", session_id)
    
    return {"message": "Sesión cerrada"}

@api_router.delete("/auth/sessions")
async def close_all_sessions(current_user: User = Depends(get_current_user), request: Request = None):
    """Cierra todas las sesiones excepto la actual"""
    current_token = None
    if request:
        auth_header = request.headers.get("authorization", "")
        if auth_header.startswith("Bearer "):
            current_token = auth_header[7:]
    
    # Obtener sesión actual por token
    current_session_id = None
    if current_token:
        current_session = await db.sessions.find_one({
            "user_id": current_user.id,
            "token_hash": hash_token(current_token),
            "is_active": True
        })
        if current_session:
            current_session_id = current_session["id"]
    
    await close_all_user_sessions(current_user.id, current_session_id)
    await log_audit(current_user.id, current_user.full_name, "all_sessions_closed", "session", None)
    
    return {"message": "Todas las demás sesiones han sido cerradas"}

# ====== USER ADMINISTRATION (ADMIN ONLY) ======

def require_admin_role(current_user: User):
    """Middleware para verificar rol de administrador"""
    if current_user.role.value not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")
    return True

@api_router.get("/admin/users")
async def get_all_users(
    search: Optional[str] = None,
    role: Optional[str] = None,
    status: Optional[str] = None,
    ambulance_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Lista todos los usuarios (solo admin/coordinador/supervisor)"""
    require_admin_role(current_user)
    
    query = {}
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    if role:
        query["role"] = role
    if status == "active":
        query["is_active"] = True
    elif status == "inactive":
        query["is_active"] = False
    if ambulance_id:
        query["assigned_ambulance_id"] = ambulance_id
    
    users = await db.users.find(query).to_list(500)
    
    result = []
    for u in users:
        # Contar sesiones activas
        sessions_count = await count_user_active_sessions(u["id"])
        
        result.append({
            "id": u["id"],
            "email": u["email"],
            "full_name": u["full_name"],
            "role": u["role"],
            "phone": u.get("phone"),
            "is_active": u.get("is_active", True),
            "assigned_ambulance_id": u.get("assigned_ambulance_id"),
            "two_fa_enabled": u.get("two_fa_enabled", False),
            "last_login": u.get("last_login"),
            "last_login_ip": u.get("last_login_ip"),
            "created_at": u["created_at"],
            "active_sessions_count": sessions_count,
            "failed_login_attempts": u.get("failed_login_attempts", 0),
            "locked_until": u.get("locked_until"),
            "password_expires_at": u.get("password_expires_at"),
        })
    
    return result

@api_router.get("/admin/users/{user_id}")
async def get_user_detail(user_id: str, current_user: User = Depends(get_current_user)):
    """Obtiene detalle completo de un usuario"""
    require_admin_role(current_user)
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Obtener sesiones activas
    sessions = await get_user_active_sessions(user_id)
    
    # Obtener historial de accesos (últimos 20)
    login_history = await db.login_attempts.find(
        {"user_id": user_id}
    ).sort("timestamp", -1).limit(20).to_list(20)
    
    # Obtener acciones críticas del usuario (últimas 50)
    audit_history = await db.audit_log.find(
        {"user_id": user_id}
    ).sort("timestamp", -1).limit(50).to_list(50)
    
    return {
        "user": {
            "id": user["id"],
            "email": user["email"],
            "full_name": user["full_name"],
            "role": user["role"],
            "phone": user.get("phone"),
            "is_active": user.get("is_active", True),
            "assigned_ambulance_id": user.get("assigned_ambulance_id"),
            "two_fa_enabled": user.get("two_fa_enabled", False),
            "two_fa_verified": user.get("two_fa_verified", False),
            "last_login": user.get("last_login"),
            "last_login_ip": user.get("last_login_ip"),
            "created_at": user["created_at"],
            "created_by": user.get("created_by"),
            "updated_at": user.get("updated_at"),
            "updated_by": user.get("updated_by"),
            "failed_login_attempts": user.get("failed_login_attempts", 0),
            "locked_until": user.get("locked_until"),
            "password_expires_at": user.get("password_expires_at"),
            "must_change_password": user.get("must_change_password", False),
        },
        "active_sessions": [{
            "id": s["id"],
            "ip_address": s.get("ip_address"),
            "device_info": s.get("device_info"),
            "created_at": s["created_at"],
            "last_activity": s.get("last_activity"),
        } for s in sessions],
        "login_history": [{
            "success": l["success"],
            "ip_address": l.get("ip_address"),
            "failure_reason": l.get("failure_reason"),
            "timestamp": l["timestamp"],
        } for l in login_history],
        "audit_history": [{
            "action": a["action"],
            "target_type": a.get("target_type"),
            "target_id": a.get("target_id"),
            "details": a.get("details"),
            "timestamp": a["timestamp"],
        } for a in audit_history],
    }

@api_router.post("/admin/users")
async def create_user_admin(data: UserCreateAdmin, current_user: User = Depends(get_current_user)):
    """Crea un nuevo usuario (solo admin)"""
    require_admin_role(current_user)
    
    # Verificar que no existe
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    # Validar contraseña
    valid, msg = validate_password_strength(data.password)
    if not valid:
        raise HTTPException(status_code=400, detail=msg)
    
    # Verificar límite de usuarios
    config = await get_system_config()
    active_count = await count_active_users()
    max_users = config.get("max_active_users", SystemConfig.MAX_ACTIVE_USERS)
    
    if data.is_active and active_count >= max_users:
        raise HTTPException(status_code=400, detail=f"Se ha alcanzado el límite de {max_users} usuarios activos")
    
    # Notificar si está cerca del límite (90%)
    if active_count >= max_users * 0.9:
        await notify_all_coordinators(
            "⚠️ Límite de usuarios al 90%",
            f"Se han utilizado {active_count}/{max_users} usuarios activos",
            NotificationType.ALERTA_INVENTARIO,  # Reutilizar tipo
            NotificationPriority.ALERTA
        )
    
    user_id = str(uuid.uuid4())
    now = datetime.utcnow()
    
    requires_2fa = data.role.value in TWO_FA_REQUIRED_ROLES
    
    user_dict = {
        "id": user_id,
        "email": data.email,
        "full_name": data.full_name,
        "role": data.role.value,
        "phone": data.phone,
        "hashed_password": hash_password(data.password),
        "password_history": [{"password_hash": hash_password(data.password), "created_at": now}],
        "password_expires_at": calculate_password_expiry(),
        "must_change_password": True,  # Forzar cambio en primer login
        "created_at": now,
        "created_by": current_user.id,
        "is_active": data.is_active,
        "two_fa_enabled": False,
        "two_fa_secret": None,
        "two_fa_verified": False,
        "two_fa_required": requires_2fa,
        "failed_login_attempts": 0,
        "locked_until": None,
        "assigned_ambulance_id": data.assigned_ambulance_id,
    }
    
    await db.users.insert_one(user_dict)
    await log_audit(current_user.id, current_user.full_name, "user_created", "user", user_id, 
                   {"email": data.email, "role": data.role.value})
    
    return {"message": "Usuario creado correctamente", "user_id": user_id}

@api_router.put("/admin/users/{user_id}")
async def update_user_admin(user_id: str, data: UserUpdate, current_user: User = Depends(get_current_user)):
    """Actualiza un usuario (solo admin)"""
    require_admin_role(current_user)
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # No permitir que un usuario cambie su propio rol
    if user_id == current_user.id and data.role and data.role.value != user["role"]:
        raise HTTPException(status_code=403, detail="No puede cambiar su propio rol")
    
    update_data = {"updated_at": datetime.utcnow(), "updated_by": current_user.id}
    audit_details = {}
    
    if data.email and data.email != user["email"]:
        # Verificar que no exista otro usuario con ese email
        existing = await db.users.find_one({"email": data.email, "id": {"$ne": user_id}})
        if existing:
            raise HTTPException(status_code=400, detail="El email ya está en uso")
        update_data["email"] = data.email
        audit_details["email_changed"] = {"from": user["email"], "to": data.email}
    
    if data.full_name:
        update_data["full_name"] = data.full_name
        audit_details["full_name"] = data.full_name
    
    if data.role and data.role.value != user["role"]:
        update_data["role"] = data.role.value
        audit_details["role_changed"] = {"from": user["role"], "to": data.role.value}
        
        # Si el nuevo rol requiere 2FA, marcarlo
        if data.role.value in TWO_FA_REQUIRED_ROLES:
            update_data["two_fa_required"] = True
    
    if data.phone is not None:
        update_data["phone"] = data.phone
    
    if data.assigned_ambulance_id is not None:
        update_data["assigned_ambulance_id"] = data.assigned_ambulance_id
        audit_details["ambulance_assigned"] = data.assigned_ambulance_id
    
    if data.is_active is not None and data.is_active != user.get("is_active", True):
        update_data["is_active"] = data.is_active
        action = "user_activated" if data.is_active else "user_deactivated"
        audit_details["status_changed"] = data.is_active
        
        # Si se desactiva, cerrar todas las sesiones
        if not data.is_active:
            await close_all_user_sessions(user_id)
    
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    await log_audit(current_user.id, current_user.full_name, "user_updated", "user", user_id, audit_details)
    
    return {"message": "Usuario actualizado correctamente"}

@api_router.post("/admin/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, data: PasswordReset, current_user: User = Depends(get_current_user)):
    """Resetea la contraseña de un usuario (solo admin)"""
    require_admin_role(current_user)
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Validar nueva contraseña
    valid, msg = validate_password_strength(data.new_password)
    if not valid:
        raise HTTPException(status_code=400, detail=msg)
    
    new_hash = hash_password(data.new_password)
    password_history = user.get("password_history", [])
    password_history.append({"password_hash": new_hash, "created_at": datetime.utcnow()})
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "hashed_password": new_hash,
            "password_history": password_history[-SystemConfig.PASSWORD_HISTORY_COUNT:],
            "password_expires_at": calculate_password_expiry(),
            "must_change_password": True,
            "failed_login_attempts": 0,
            "locked_until": None,
            "updated_at": datetime.utcnow(),
            "updated_by": current_user.id
        }}
    )
    
    # Cerrar todas las sesiones del usuario
    await close_all_user_sessions(user_id)
    
    await log_audit(current_user.id, current_user.full_name, "password_reset", "user", user_id)
    
    return {"message": "Contraseña reseteada. El usuario deberá cambiarla en su próximo login."}

@api_router.post("/admin/users/{user_id}/unlock")
async def unlock_user(user_id: str, current_user: User = Depends(get_current_user)):
    """Desbloquea un usuario bloqueado por intentos fallidos"""
    require_admin_role(current_user)
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "failed_login_attempts": 0,
            "locked_until": None,
            "updated_at": datetime.utcnow(),
            "updated_by": current_user.id
        }}
    )
    
    await log_audit(current_user.id, current_user.full_name, "user_unlocked", "user", user_id)
    
    return {"message": "Usuario desbloqueado"}

@api_router.delete("/admin/users/{user_id}/sessions/{session_id}")
async def close_user_session_admin(user_id: str, session_id: str, current_user: User = Depends(get_current_user)):
    """Cierra una sesión específica de un usuario (admin)"""
    require_admin_role(current_user)
    
    session = await db.sessions.find_one({"id": session_id, "user_id": user_id})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    
    await close_session(session_id)
    await log_audit(current_user.id, current_user.full_name, "session_closed_remote", "session", session_id,
                   {"target_user_id": user_id})
    
    return {"message": "Sesión cerrada"}

@api_router.delete("/admin/users/{user_id}/sessions")
async def close_all_user_sessions_admin(user_id: str, current_user: User = Depends(get_current_user)):
    """Cierra todas las sesiones de un usuario (admin)"""
    require_admin_role(current_user)
    
    await close_all_user_sessions(user_id)
    await log_audit(current_user.id, current_user.full_name, "all_sessions_closed", "session", None,
                   {"target_user_id": user_id})
    
    return {"message": "Todas las sesiones del usuario han sido cerradas"}

# ====== SYSTEM CONFIG ======

@api_router.get("/admin/config")
async def get_config(current_user: User = Depends(get_current_user)):
    """Obtiene la configuración del sistema"""
    if current_user.role.value != UserRole.ADMINISTRADOR.value:
        raise HTTPException(status_code=403, detail="Solo el administrador puede ver la configuración")
    
    config = await get_system_config()
    active_users = await count_active_users()
    
    return {
        "max_active_users": config.get("max_active_users", SystemConfig.MAX_ACTIVE_USERS),
        "current_active_users": active_users,
        "password_expiry_days": config.get("password_expiry_days", SystemConfig.PASSWORD_EXPIRY_DAYS),
        "two_fa_optional_for_paramedics": config.get("two_fa_optional_for_paramedics", True),
    }

@api_router.put("/admin/config")
async def update_config(
    max_active_users: Optional[int] = None,
    password_expiry_days: Optional[int] = None,
    two_fa_optional_for_paramedics: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    """Actualiza la configuración del sistema (solo administrador)"""
    if current_user.role.value != UserRole.ADMINISTRADOR.value:
        raise HTTPException(status_code=403, detail="Solo el administrador puede modificar la configuración")
    
    update_data = {"updated_at": datetime.utcnow(), "updated_by": current_user.id}
    
    if max_active_users is not None:
        if max_active_users < 1:
            raise HTTPException(status_code=400, detail="El límite debe ser al menos 1")
        update_data["max_active_users"] = max_active_users
    
    if password_expiry_days is not None:
        if password_expiry_days < 30 or password_expiry_days > 365:
            raise HTTPException(status_code=400, detail="Los días de expiración deben estar entre 30 y 365")
        update_data["password_expiry_days"] = password_expiry_days
    
    if two_fa_optional_for_paramedics is not None:
        update_data["two_fa_optional_for_paramedics"] = two_fa_optional_for_paramedics
    
    await db.system_config.update_one(
        {"id": "main"},
        {"$set": update_data},
        upsert=True
    )
    
    await log_audit(current_user.id, current_user.full_name, "config_updated", "system", "main", update_data)
    
    return {"message": "Configuración actualizada"}

# ====== ROLE-BASED MIDDLEWARE FOR FINANCE ======

def require_finance_access(current_user: User):
    """Middleware para verificar acceso a finanzas"""
    if current_user.role.value not in FULL_ACCESS_ROLES:
        # Registrar intento de acceso denegado
        raise HTTPException(status_code=403, detail="No tiene acceso al módulo financiero")
    return True

# Ambulance Routes
@api_router.post("/ambulances", response_model=Ambulance)
async def create_ambulance(data: AmbulanceCreate, current_user: User = Depends(get_current_user)):
    ambulance_dict = {
        "id": str(uuid.uuid4()),
        **data.dict(),
        "created_at": datetime.utcnow(),
        "services_count": 0
    }
    await db.ambulances.insert_one(ambulance_dict)
    await log_audit(current_user.id, current_user.full_name, "crear", "ambulancia", ambulance_dict["id"], {"unit_number": data.unit_number})
    return Ambulance(**ambulance_dict)

@api_router.get("/ambulances", response_model=List[Ambulance])
async def get_ambulances(status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    ambulances = await db.ambulances.find(query).to_list(100)
    return [Ambulance(**a) for a in ambulances]

@api_router.get("/ambulances/{ambulance_id}", response_model=Ambulance)
async def get_ambulance(ambulance_id: str, current_user: User = Depends(get_current_user)):
    ambulance = await db.ambulances.find_one({"id": ambulance_id})
    if not ambulance:
        raise HTTPException(status_code=404, detail="Ambulancia no encontrada")
    return Ambulance(**ambulance)

@api_router.put("/ambulances/{ambulance_id}", response_model=Ambulance)
async def update_ambulance(ambulance_id: str, data: AmbulanceCreate, current_user: User = Depends(get_current_user)):
    result = await db.ambulances.update_one(
        {"id": ambulance_id},
        {"$set": data.dict()}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Ambulancia no encontrada")
    ambulance = await db.ambulances.find_one({"id": ambulance_id})
    await log_audit(current_user.id, current_user.full_name, "actualizar", "ambulancia", ambulance_id, data.dict())
    return Ambulance(**ambulance)

@api_router.delete("/ambulances/{ambulance_id}")
async def delete_ambulance(ambulance_id: str, current_user: User = Depends(get_current_user)):
    result = await db.ambulances.delete_one({"id": ambulance_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ambulancia no encontrada")
    await log_audit(current_user.id, current_user.full_name, "eliminar", "ambulancia", ambulance_id, {})
    return {"message": "Ambulancia eliminada"}

# Personnel Routes
@api_router.post("/personnel", response_model=Personnel)
async def create_personnel(data: PersonnelCreate, current_user: User = Depends(get_current_user)):
    personnel_dict = {
        "id": str(uuid.uuid4()),
        **data.dict(),
        "created_at": datetime.utcnow(),
        "services_count": 0
    }
    await db.personnel.insert_one(personnel_dict)
    await log_audit(current_user.id, current_user.full_name, "crear", "personal", personnel_dict["id"], {"full_name": data.full_name})
    return Personnel(**personnel_dict)

@api_router.get("/personnel", response_model=List[Personnel])
async def get_personnel(is_available: Optional[bool] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if is_available is not None:
        query["is_available"] = is_available
    personnel = await db.personnel.find(query).to_list(100)
    return [Personnel(**p) for p in personnel]

@api_router.get("/personnel/{personnel_id}", response_model=Personnel)
async def get_personnel_by_id(personnel_id: str, current_user: User = Depends(get_current_user)):
    person = await db.personnel.find_one({"id": personnel_id})
    if not person:
        raise HTTPException(status_code=404, detail="Personal no encontrado")
    return Personnel(**person)

@api_router.put("/personnel/{personnel_id}", response_model=Personnel)
async def update_personnel(personnel_id: str, data: PersonnelCreate, current_user: User = Depends(get_current_user)):
    result = await db.personnel.update_one(
        {"id": personnel_id},
        {"$set": data.dict()}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Personal no encontrado")
    personnel = await db.personnel.find_one({"id": personnel_id})
    await log_audit(current_user.id, current_user.full_name, "actualizar", "personal", personnel_id, data.dict())
    return Personnel(**personnel)

@api_router.delete("/personnel/{personnel_id}")
async def delete_personnel(personnel_id: str, current_user: User = Depends(get_current_user)):
    result = await db.personnel.delete_one({"id": personnel_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Personal no encontrado")
    await log_audit(current_user.id, current_user.full_name, "eliminar", "personal", personnel_id, {})
    return {"message": "Personal eliminado"}

# Service Routes
@api_router.post("/services", response_model=Service)
async def create_service(data: ServiceCreate, current_user: User = Depends(get_current_user)):
    # Check available ambulances
    await check_available_ambulances()
    
    # If equipment requires oxygen or ventilator, validate availability
    if data.equipment_required:
        if data.equipment_required.oxygen and not data.equipment_required.oxygen_liters:
            raise HTTPException(status_code=400, detail="Si se requiere oxígeno, debe especificar los litros requeridos")
        if data.equipment_required.infusion_pumps and not data.equipment_required.infusion_pumps_count:
            raise HTTPException(status_code=400, detail="Si se requieren bombas de infusión, debe especificar la cantidad")
    
    # If personnel requires doctor, validate doctor name
    if data.personnel_required and data.personnel_required.doctor and not data.personnel_required.doctor_name:
        raise HTTPException(status_code=400, detail="Si se requiere médico, debe especificar el nombre")
    
    service_dict = {
        "id": str(uuid.uuid4()),
        **data.dict(),
        "status": ServiceStatus.PENDIENTE,
        "ambulance_id": None,
        "personnel_ids": [],
        "payment_type": PaymentType.PENDIENTE,
        "payment_amount": None,
        "frap_notes": None,
        "bank_name": None,
        "account_number": None,
        "account_holder": None,
        "transfer_reference": None,
        "payment_proof_base64": None,
        "cost_adjustments": [],  # Para historial de ajustes
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow(),
        "created_by": current_user.id,
        "status_history": [{
            "status": "pendiente",
            "timestamp": datetime.utcnow().isoformat(),
            "user_id": current_user.id,
            "user_name": current_user.full_name
        }]
    }
    await db.services.insert_one(service_dict)
    
    # Determine notification type and priority
    notification_type = NotificationType.URGENCIA if data.service_type == ServiceType.URGENTE else NotificationType.SERVICIO_NUEVO
    priority = NotificationPriority.NORMAL
    emoji = "📋"
    
    if data.service_type == ServiceType.URGENTE:
        priority = NotificationPriority.ALERTA
        emoji = "🚨"
    
    # If patient status is CRITICAL, send critical notification
    if data.patient_status == PatientStatus.CRITICO:
        priority = NotificationPriority.CRITICA
        emoji = "🚨🚨"
        await notify_all_coordinators(
            f"{emoji} PACIENTE CRÍTICO - Servicio urgente",
            f"Paciente: {data.patient.name} - Estado: CRÍTICO - De: {data.origin} A: {data.destination}",
            NotificationType.URGENCIA,
            priority,
            service_dict["id"]
        )
    else:
        await notify_all_coordinators(
            f"{emoji} Nuevo servicio {data.service_type.value}",
            f"Paciente: {data.patient.name} - De: {data.origin} A: {data.destination}",
            notification_type,
            priority,
            service_dict["id"]
        )
    
    await log_audit(current_user.id, current_user.full_name, "crear", "servicio", service_dict["id"], 
                   {"service_type": data.service_type, "patient": data.patient.name, "patient_status": data.patient_status})
    
    return Service(**service_dict)

@api_router.get("/services", response_model=List[Service])
async def get_services(
    status: Optional[ServiceStatus] = None,
    service_type: Optional[ServiceType] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    ambulance_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if service_type:
        query["service_type"] = service_type
    if ambulance_id:
        query["ambulance_id"] = ambulance_id
    if date_from:
        query["scheduled_date"] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        if "scheduled_date" in query:
            query["scheduled_date"]["$lte"] = datetime.fromisoformat(date_to)
        else:
            query["scheduled_date"] = {"$lte": datetime.fromisoformat(date_to)}
    
    services = await db.services.find(query).sort("scheduled_date", -1).to_list(500)
    return [Service(**s) for s in services]

@api_router.get("/services/active")
async def get_active_services(current_user: User = Depends(get_current_user)):
    """Get all active services (not finalized or cancelled)"""
    query = {"status": {"$in": ["pendiente", "confirmado", "en_camino", "en_curso"]}}
    services = await db.services.find(query).sort("scheduled_date", 1).to_list(100)
    return [Service(**s) for s in services]

@api_router.get("/services/{service_id}", response_model=Service)
async def get_service(service_id: str, current_user: User = Depends(get_current_user)):
    service = await db.services.find_one({"id": service_id})
    if not service:
        raise HTTPException(status_code=404, detail="Servicio no encontrado")
    return Service(**service)

@api_router.put("/services/{service_id}", response_model=Service)
async def update_service(service_id: str, data: ServiceUpdate, current_user: User = Depends(get_current_user)):
    service = await db.services.find_one({"id": service_id})
    if not service:
        raise HTTPException(status_code=404, detail="Servicio no encontrado")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    # Add status to history if changed
    if data.status and data.status != service.get("status"):
        status_history = service.get("status_history", [])
        status_history.append({
            "status": data.status,
            "timestamp": datetime.utcnow().isoformat(),
            "user_id": current_user.id,
            "user_name": current_user.full_name
        })
        update_data["status_history"] = status_history
    
    await db.services.update_one(
        {"id": service_id},
        {"$set": update_data}
    )
    
    # Update ambulance status if assigned
    if data.ambulance_id:
        await db.ambulances.update_one(
            {"id": data.ambulance_id},
            {"$set": {"status": AmbulanceStatus.EN_SERVICIO}}
        )
        await check_available_ambulances()
    
    # If service completed or cancelled, free ambulance and personnel
    if data.status in [ServiceStatus.FINALIZADO, ServiceStatus.CANCELADO]:
        if service.get("ambulance_id"):
            await db.ambulances.update_one(
                {"id": service["ambulance_id"]},
                {"$set": {"status": AmbulanceStatus.DISPONIBLE}, "$inc": {"services_count": 1}}
            )
        for p_id in service.get("personnel_ids", []):
            await db.personnel.update_one(
                {"id": p_id},
                {"$set": {"is_available": True}, "$inc": {"services_count": 1}}
            )
    
    # Mark personnel as unavailable when assigned
    if data.personnel_ids:
        for p_id in data.personnel_ids:
            await db.personnel.update_one(
                {"id": p_id},
                {"$set": {"is_available": False}}
            )
    
    # Create finance entry if payment received
    if data.payment_amount and data.payment_type and data.payment_type != PaymentType.PENDIENTE:
        finance_entry = {
            "id": str(uuid.uuid4()),
            "entry_type": "ingreso",
            "amount": data.payment_amount,
            "category": "servicio",
            "description": f"Pago por servicio - {service.get('patient', {}).get('name', 'N/A')}",
            "payment_type": data.payment_type,
            "service_id": service_id,
            "bank_name": data.bank_name,
            "account_number": data.account_number,
            "account_holder": data.account_holder,
            "transfer_reference": data.transfer_reference,
            "receipt_base64": data.payment_proof_base64,
            "created_at": datetime.utcnow(),
            "created_by": current_user.id
        }
        await db.finances.insert_one(finance_entry)
    
    updated_service = await db.services.find_one({"id": service_id})
    await log_audit(current_user.id, current_user.full_name, "actualizar", "servicio", service_id, update_data)
    
    return Service(**updated_service)

@api_router.delete("/services/{service_id}")
async def delete_service(service_id: str, current_user: User = Depends(get_current_user)):
    result = await db.services.delete_one({"id": service_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Servicio no encontrado")
    await log_audit(current_user.id, current_user.full_name, "eliminar", "servicio", service_id, {})
    return {"message": "Servicio eliminado"}

# Notifications Routes
@api_router.get("/notifications", response_model=List[Notification])
async def get_notifications(current_user: User = Depends(get_current_user)):
    notifications = await db.notifications.find({"user_id": current_user.id}).sort("created_at", -1).to_list(100)
    return [Notification(**n) for n in notifications]

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: User = Depends(get_current_user)):
    await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user.id},
        {"$set": {"is_read": True}}
    )
    return {"message": "Notificación marcada como leída"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(current_user: User = Depends(get_current_user)):
    await db.notifications.update_many(
        {"user_id": current_user.id},
        {"$set": {"is_read": True}}
    )
    return {"message": "Todas las notificaciones marcadas como leídas"}

# Device Token Routes (Preparado para FCM)
class DeviceTokenCreate(BaseModel):
    token: str
    device_type: str  # "android", "ios", "web"

@api_router.post("/device-tokens")
async def register_device_token(data: DeviceTokenCreate, current_user: User = Depends(get_current_user)):
    """Registrar token de dispositivo para notificaciones push"""
    # Check if token already exists for this user
    existing = await db.device_tokens.find_one({
        "user_id": current_user.id,
        "token": data.token
    })
    
    if existing:
        # Update existing token
        await db.device_tokens.update_one(
            {"id": existing["id"]},
            {"$set": {"is_active": True, "updated_at": datetime.utcnow()}}
        )
        return {"message": "Token actualizado", "id": existing["id"]}
    
    # Create new token
    token_dict = {
        "id": str(uuid.uuid4()),
        "user_id": current_user.id,
        "token": data.token,
        "device_type": data.device_type,
        "is_active": True,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    await db.device_tokens.insert_one(token_dict)
    return {"message": "Token registrado", "id": token_dict["id"]}

@api_router.delete("/device-tokens/{token}")
async def remove_device_token(token: str, current_user: User = Depends(get_current_user)):
    """Desactivar token de dispositivo"""
    await db.device_tokens.update_one(
        {"user_id": current_user.id, "token": token},
        {"$set": {"is_active": False, "updated_at": datetime.utcnow()}}
    )
    return {"message": "Token desactivado"}
@api_router.post("/users/fcm-token")
async def update_fcm_token(data: dict, current_user: User = Depends(get_current_user)):
    """Registrar token FCM del dispositivo para notificaciones push"""
    fcm_token = data.get("fcm_token")
    if not fcm_token:
        raise HTTPException(status_code=400, detail="Token FCM requerido")
    
    existing = await db.device_tokens.find_one({
        "user_id": current_user.id,
        "token": fcm_token
    })
    
    if existing:
        await db.device_tokens.update_one(
            {"user_id": current_user.id, "token": fcm_token},
            {"$set": {"is_active": True, "updated_at": datetime.utcnow()}}
        )
    else:
        token_dict = {
            "id": str(uuid.uuid4()),
            "user_id": current_user.id,
            "token": fcm_token,
            "device_type": "android",
            "is_active": True,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        await db.device_tokens.insert_one(token_dict)
    
    return {"message": "Token FCM registrado correctamente"}
# Finance Routes
@api_router.post("/finances", response_model=FinanceEntry)
async def create_finance_entry(data: FinanceEntryCreate, current_user: User = Depends(get_current_user)):
    # Verificar acceso a finanzas
    require_finance_access(current_user)
    
    entry_dict = {
        "id": str(uuid.uuid4()),
        **data.dict(),
        "created_at": datetime.utcnow(),
        "created_by": current_user.id
    }
    await db.finances.insert_one(entry_dict)
    await log_audit(current_user.id, current_user.full_name, "finance_created", "finanza", entry_dict["id"], 
                   {"entry_type": data.entry_type, "amount": data.amount})
    return FinanceEntry(**entry_dict)

@api_router.get("/finances")
async def get_finances(
    entry_type: Optional[str] = None,
    category: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Verificar acceso a finanzas
    require_finance_access(current_user)
    
    # Registrar acceso en auditoría
    await log_audit(current_user.id, current_user.full_name, "finance_viewed", "finanzas", None)
    
    query = {}
    if entry_type:
        query["entry_type"] = entry_type
    if category:
        query["category"] = category
    if date_from:
        query["created_at"] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(date_to)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(date_to)}
    
    entries = await db.finances.find(query).sort("created_at", -1).to_list(500)
    return [FinanceEntry(**e) for e in entries]

@api_router.get("/finances/summary")
async def get_finance_summary(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Verificar acceso a finanzas
    require_finance_access(current_user)
    query = {}
    if date_from:
        query["created_at"] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(date_to)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(date_to)}
    
    entries = await db.finances.find(query).to_list(1000)
    
    total_ingresos = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso")
    total_egresos = sum(e["amount"] for e in entries if e["entry_type"] == "egreso")
    
    ingresos_efectivo = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "efectivo")
    ingresos_transferencia = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "transferencia")
    
    egresos_by_category = {}
    for e in entries:
        if e["entry_type"] == "egreso":
            cat = e.get("category", "otros")
            egresos_by_category[cat] = egresos_by_category.get(cat, 0) + e["amount"]
    
    return {
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "balance": total_ingresos - total_egresos,
        "ingresos_efectivo": ingresos_efectivo,
        "ingresos_transferencia": ingresos_transferencia,
        "egresos_by_category": egresos_by_category
    }

# Checklist Routes
@api_router.post("/checklists")
async def create_checklist(data: ChecklistCreate, current_user: User = Depends(get_current_user)):
    # Count failures - handle both dict format and object format
    all_items = (
        data.apariencia_general + data.cabina_operadores + data.compartimiento_motor +
        data.exterior_operador + data.zona_frontal + data.exterior_copiloto +
        data.compartimento_paciente + data.zona_posterior + data.herramientas
    )
    
    # Handle both old and new format
    failure_count = sum(1 for item in all_items if not (item.get('status', True) if isinstance(item, dict) else item.status))
    has_failures = failure_count > 0
    
    # Check for critical failures (items that could affect safety)
    critical_items = ["Frenos", "Luces de emergencia", "Sirena", "Extintor", "Batería", "Fugas", "Asientos y cinturones", "Llantas"]
    critical_failures = []
    for item in all_items:
        item_status = item.get('status', True) if isinstance(item, dict) else item.status
        item_name = item.get('name', '') if isinstance(item, dict) else item.name
        item_critical = item.get('critical', False) if isinstance(item, dict) else getattr(item, 'critical', False)
        
        if not item_status and (item_name in critical_items or item_critical):
            critical_failures.append(item_name)
    
    # Check nivel críticos
    for nivel in data.niveles:
        nivel_name = nivel.get('name', '') if isinstance(nivel, dict) else nivel.name
        nivel_level = nivel.get('level', 'normal') if isinstance(nivel, dict) else nivel.level
        nivel_critical = nivel.get('critical', False) if isinstance(nivel, dict) else getattr(nivel, 'critical', False)
        
        if nivel_level == 'bajo' and nivel_critical:
            critical_failures.append(f"{nivel_name} (bajo)")
            failure_count += 1
            has_failures = True
    
    checklist_dict = {
        "id": str(uuid.uuid4()),
        **data.dict(),
        "created_at": datetime.utcnow(),
        "has_failures": has_failures,
        "failure_count": data.failure_count if data.failure_count else failure_count,
        "critical_failures": critical_failures,
        "has_critical_failures": len(critical_failures) > 0,
        "signed": data.signature_base64 is not None,
        "signed_by": current_user.full_name,
        "signed_at": data.completed_at or datetime.utcnow()
    }
    await db.checklists.insert_one(checklist_dict)
    
    ambulance = await db.ambulances.find_one({"id": data.ambulance_id})
    unit_number = ambulance.get('unit_number', 'N/A') if ambulance else 'N/A'
    
    # If there are critical failures, mark ambulance as out of service and notify with CRITICAL priority
    if len(critical_failures) > 0:
        await db.ambulances.update_one(
            {"id": data.ambulance_id},
            {"$set": {"status": AmbulanceStatus.FUERA_SERVICIO}}
        )
        await notify_all_coordinators(
            f"🚨 CRÍTICO: Unidad {unit_number} fuera de servicio",
            f"Fallas críticas: {', '.join(critical_failures)}. La unidad ha sido deshabilitada.",
            NotificationType.UNIDAD_FUERA_SERVICIO,
            NotificationPriority.CRITICA,
            entity_id=data.ambulance_id,
            entity_type="ambulancia"
        )
    elif has_failures:
        # Non-critical failures - alert priority
        await notify_all_coordinators(
            f"⚠️ Fallas en checklist - Unidad {unit_number}",
            f"Se detectaron {failure_count} fallas en el checklist. Revisar observaciones.",
            NotificationType.ALERTA_CHECKLIST,
            NotificationPriority.ALERTA,
            entity_id=data.ambulance_id,
            entity_type="ambulancia"
        )
    
    # Check if there are critical observations
    if data.observations and any(word in data.observations.lower() for word in ['urgente', 'crítico', 'peligro', 'fuga', 'humo']):
        await notify_all_coordinators(
            f"⚠️ Observación crítica - Unidad {unit_number}",
            f"Observaciones: {data.observations[:100]}...",
            NotificationType.ALERTA_CHECKLIST,
            NotificationPriority.CRITICA,
            entity_id=data.ambulance_id,
            entity_type="ambulancia"
        )
    
    await log_audit(current_user.id, current_user.full_name, "crear", "checklist", checklist_dict["id"],
                   {"ambulance_id": data.ambulance_id, "failures": failure_count, "critical": len(critical_failures), "signed": data.signature_base64 is not None})
    
    # Remove MongoDB's _id before returning
    checklist_dict.pop("_id", None)
    return checklist_dict

@api_router.get("/checklists")
async def get_checklists(
    ambulance_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    has_failures: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if ambulance_id:
        query["ambulance_id"] = ambulance_id
    if has_failures is not None:
        query["has_failures"] = has_failures
    if date_from:
        query["date"] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        if "date" in query:
            query["date"]["$lte"] = datetime.fromisoformat(date_to)
        else:
            query["date"] = {"$lte": datetime.fromisoformat(date_to)}
    
    checklists = await db.checklists.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return checklists

@api_router.get("/checklists/{checklist_id}")
async def get_checklist(checklist_id: str, current_user: User = Depends(get_current_user)):
    checklist = await db.checklists.find_one({"id": checklist_id}, {"_id": 0})
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist no encontrado")
    return checklist

# ==========================================
# INVENTORY ROUTES - Sistema por Ambulancia
# ==========================================

# Definición de items predeterminados por categoría
INVENTORY_TEMPLATE = {
    InventoryCategory.BOTIQUIN_TRAUMA: [
        {"name": "Abatelenguas", "expected": 5},
        {"name": "Apósitos", "expected": 5},
        {"name": "Gasas 7.5x5", "expected": 5},
        {"name": "Gasas 10x10", "expected": 5},
        {"name": "Gasas 10x10 estériles", "expected": 5},
        {"name": "Pañales / Protectores", "expected": 2},
        {"name": "Sábana térmica", "expected": 1},
        {"name": "Parches / Electrodos", "expected": 20},
        {"name": "Venda elástica 5cm", "expected": 2},
        {"name": "Venda elástica 10cm", "expected": 2},
        {"name": "Venda elástica 15cm", "expected": 2},
        {"name": "Venda elástica 20cm", "expected": 2},
        {"name": "Cabestrillo", "expected": 1},
        {"name": "Equipo de venoclisis", "expected": 2},
        {"name": "Ligadura / Torniquete", "expected": 1},
        {"name": "Punzocat #14", "expected": 2},
        {"name": "Punzocat #16", "expected": 2},
        {"name": "Punzocat #18", "expected": 2},
        {"name": "Punzocat #20", "expected": 2},
        {"name": "Punzocat #22", "expected": 2},
        {"name": "Punzocat #24", "expected": 2},
        {"name": "Sol. Hartmann 500ml", "expected": 1},
        {"name": "Sol. Glucosa 5% 500ml", "expected": 1},
        {"name": "Sol. Glucosa 50% 50ml", "expected": 1},
        {"name": "Sol. NaCl 0.9% 500ml", "expected": 1},
        {"name": "Tela adhesiva y micropore", "expected": 1},
        {"name": "Torundas con alcohol", "expected": 1},
        {"name": "Jeringa 1ml", "expected": 3},
        {"name": "Jeringa 3ml", "expected": 3},
        {"name": "Jeringa 5ml", "expected": 3},
        {"name": "Jeringa 10ml", "expected": 3},
        {"name": "Jeringa 20ml", "expected": 3},
        {"name": "Jeringa de insulina 1ml", "expected": 1},
        {"name": "Jeringas de insulina 1ml", "expected": 3},
        {"name": "Baumanómetro y estetoscopio adulto", "expected": 1},
        {"name": "Termómetro digital", "expected": 1},
        {"name": "Oxímetro de pulso", "expected": 1},
        {"name": "Lámpara diagnóstica", "expected": 1},
        {"name": "Glucómetro", "expected": 1},
        {"name": "Tiras reactivas y lancetas", "expected": 15},
        {"name": "Tijeras uso rudo", "expected": 1},
    ],
    InventoryCategory.BOTIQUIN_VIA_AEREA: [
        {"name": "BVM adulto", "expected": 1},
        {"name": "BVM pediátrico", "expected": 1},
        {"name": "BVM neonatal", "expected": 1},
        {"name": "Cánulas orofaríngeas Berman", "expected": 3},
        {"name": "Cánulas orofaríngeas Guedel", "expected": 3},
        {"name": "Gasas 10x10 estériles", "expected": 3},
        {"name": "Abatelenguas", "expected": 5},
        {"name": "Mascarilla con reservorio adulto", "expected": 2},
        {"name": "Mascarilla con reservorio pediátrica", "expected": 2},
        {"name": "Puntas nasales adulto", "expected": 2},
        {"name": "Puntas nasales pediátricas", "expected": 2},
        {"name": "Línea de aspiración", "expected": 2},
        {"name": "Cánula Yankauer", "expected": 2},
        {"name": "Cánula French", "expected": 2},
    ],
    InventoryCategory.MEDICAMENTOS: [
        {"name": "Ácido acetilsalicílico", "expected": 1, "requires_expiry": True},
        {"name": "Dinitrato de isosorbida", "expected": 1, "requires_expiry": True},
        {"name": "Salbutamol aerosol", "expected": 1, "requires_expiry": True},
        {"name": "Ketorolaco IV/IM", "expected": 2, "requires_expiry": True},
        {"name": "Metamizol sódico IV/IM", "expected": 2, "requires_expiry": True},
        {"name": "Epinefrina IV/IM", "expected": 4, "requires_expiry": True, "is_critical": True},
        {"name": "Dexametasona IV/IM", "expected": 2, "requires_expiry": True},
        {"name": "Atropina", "expected": 2, "requires_expiry": True, "is_critical": True},
        {"name": "Electrolitos orales", "expected": 8, "requires_expiry": True},
    ],
    InventoryCategory.RACK: [
        {"name": "Apósitos", "expected": 5},
        {"name": "Gasas 7.5x5", "expected": 3},
        {"name": "Gasas 10x10", "expected": 5},
        {"name": "Frasco torundas con alcohol", "expected": 1},
        {"name": "Frasco torundas con isodine", "expected": 1},
        {"name": "Jeringa 1ml", "expected": 3},
        {"name": "Jeringa 3ml", "expected": 3},
        {"name": "Jeringa 5ml", "expected": 3},
        {"name": "Jeringa 10ml", "expected": 1},
        {"name": "Jeringa 20ml", "expected": 1},
        {"name": "Pañales / protectores", "expected": 2},
        {"name": "Bolsas RPBI roja", "expected": 3},
        {"name": "Contenedor RPBI rígido", "expected": 1},
        {"name": "Equipo de venoclisis", "expected": 2},
        {"name": "Ligadura (torniquete)", "expected": 1},
        {"name": "Punzocat #14", "expected": 1},
        {"name": "Punzocat #16", "expected": 1},
        {"name": "Punzocat #18", "expected": 1},
        {"name": "Punzocat #20", "expected": 1},
        {"name": "Punzocat #22", "expected": 1},
        {"name": "Punzocat #24", "expected": 1},
        {"name": "Tela adhesiva / micropore", "expected": 1},
        {"name": "Tegaderm adulto / pediátrico", "expected": 5},
        {"name": "Sol. Hartmann 500ml", "expected": 1},
        {"name": "Sol. Glucosa 5% 500ml", "expected": 1},
        {"name": "Sol. Glucosa 50% 50ml", "expected": 1},
        {"name": "Sol. NaCl 500ml", "expected": 1},
        {"name": "Venda elástica 5cm", "expected": 1},
        {"name": "Venda elástica 10cm", "expected": 1},
        {"name": "Venda elástica 15cm", "expected": 1},
        {"name": "Venda elástica 20cm", "expected": 1},
        {"name": "Venda elástica 30cm", "expected": 1},
        {"name": "Cánula aspiración (Yankauer / French)", "expected": 2},
        {"name": "Línea de aspiración", "expected": 2},
        {"name": "Mascarilla con reservorio adulto", "expected": 2},
        {"name": "Mascarilla con reservorio pediátrica", "expected": 2},
        {"name": "Puntas nasales adulto", "expected": 2},
        {"name": "Sensor Masimo", "expected": 2},
        {"name": "Línea de infusión estándar", "expected": 1},
    ],
    InventoryCategory.CABINA_ATENCION: [
        {"name": "Barbiquejos adulto/pediátrico", "expected": 2},
        {"name": "Camilla marina", "expected": 1},
        {"name": "Carrocamilla", "expected": 1},
        {"name": "Chaleco de extricación", "expected": 1},
        {"name": "Collarín cervical rígido adulto", "expected": 2},
        {"name": "Collarín Philadelphia", "expected": 1},
        {"name": "Collarín cervical pediátrico", "expected": 2},
        {"name": "Férula espinal larga adulto", "expected": 1},
        {"name": "Férula espinal pediátrica", "expected": 1},
        {"name": "Férulas SAM", "expected": 2},
        {"name": "Kit de férulas", "expected": 4},
        {"name": "Inmovilizador de cráneo adulto", "expected": 2},
        {"name": "Inmovilizador de cráneo pediátrico", "expected": 2},
        {"name": "Rodillo", "expected": 1},
        {"name": "Aspirador secreciones portátil - Batería", "expected": 1},
        {"name": "Aspirador secreciones portátil - Cánula Yankauer", "expected": 1},
        {"name": "Aspirador secreciones portátil - Cánula French", "expected": 1},
        {"name": "Toallas ginecológicas", "expected": 2},
    ],
    InventoryCategory.EQUIPO_PORTATIL: [
        {"name": "Aspirador de secreciones portátil", "expected": 1, "is_critical": True},
        {"name": "Batería", "expected": 1},
        {"name": "Cánula Yankauer", "expected": 1},
        {"name": "Cánula French", "expected": 1},
        {"name": "Manguera de oxígeno", "expected": 1},
    ],
    InventoryCategory.MONITOR_SIGNOS_VITALES: [
        {"name": "Pulsioxímetro adulto", "expected": 1, "is_critical": True},
        {"name": "Pulsioxímetro pediátrico", "expected": 1, "is_critical": True},
        {"name": "Baumanómetro adulto", "expected": 1},
        {"name": "Baumanómetro pediátrico", "expected": 1},
        {"name": "Termómetro", "expected": 1},
        {"name": "Electrodos", "expected": 10},
        {"name": "Desfibrilador", "expected": 1, "is_critical": True},
        {"name": "Ventilador", "expected": 1, "is_critical": True},
    ],
   InventoryCategory.OTROS: [
        {"name": "Cómodo", "expected": 1},
        {"name": "Pato urinario", "expected": 1},
        {"name": "Sábanas para carrocamilla", "expected": 4},
        {"name": "Cobertor STEN", "expected": 1},
        {"name": "Almohada c/funda", "expected": 1},
        {"name": "Gel antibacterial", "expected": 1},
        {"name": "Cubrebocas", "expected": 20},
        {"name": "Guantes látex", "expected": 20},
        {"name": "Toallas ginecológicas", "expected": 2},
        {"name": "Sabana térmica", "expected": 1},
    ],
}

# Función para calcular litros desde PSI
def calculate_liters_from_psi(current_psi: int, max_psi: int, capacity_liters: int) -> int:
    """Calcular litros restantes basado en PSI actual"""
    if max_psi == 0:
        return 0
    ratio = current_psi / max_psi
    return int(capacity_liters * ratio)

# Función para determinar estado del tanque
def get_oxygen_tank_status(liters: int) -> OxygenTankStatus:
    """Determinar estado semáforo del tanque"""
    if liters <= 0:
        return OxygenTankStatus.CRITICO
    elif liters <= 200:
        return OxygenTankStatus.ROJO
    elif liters <= 800:
        return OxygenTankStatus.AMARILLO
    else:
        return OxygenTankStatus.VERDE

# ==========================================
# OXYGEN TANK ROUTES
# ==========================================

@api_router.post("/oxygen-tanks")
async def create_oxygen_tank(data: OxygenTankCreate, current_user: User = Depends(get_current_user)):
    """Crear un nuevo tanque de oxígeno"""
    current_liters = calculate_liters_from_psi(data.current_psi, 2000, data.capacity_liters)
    status = get_oxygen_tank_status(current_liters)
    
    tank_dict = {
        "id": str(uuid.uuid4()),
        "ambulance_id": data.ambulance_id,
        "tank_type": data.tank_type,
        "is_portable": data.is_portable,
        "capacity_liters": data.capacity_liters,
        "max_psi": 2000,
        "current_psi": data.current_psi,
        "current_liters": current_liters,
        "status": status,
        "last_refill_date": None,
        "last_movement_by": current_user.full_name,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    await db.oxygen_tanks.insert_one(tank_dict)
    
    # Check for critical status
    if status in [OxygenTankStatus.ROJO, OxygenTankStatus.CRITICO]:
        ambulance = await db.ambulances.find_one({"id": data.ambulance_id})
        unit = ambulance.get("unit_number", "N/A") if ambulance else "N/A"
        priority = NotificationPriority.CRITICA if status == OxygenTankStatus.CRITICO else NotificationPriority.ALERTA
        await notify_all_coordinators(
            f"🚨 OXÍGENO {'CRÍTICO' if status == OxygenTankStatus.CRITICO else 'BAJO'} - Unidad {unit}",
            f"Tanque tipo {data.tank_type}: {current_liters}L restantes ({data.current_psi} PSI)",
            NotificationType.ALERTA_OXIGENO,
            priority,
            entity_id=tank_dict["id"],
            entity_type="oxigeno"
        )
    
    await log_audit(current_user.id, current_user.full_name, "crear", "tanque_oxigeno", tank_dict["id"],
                   {"ambulance_id": data.ambulance_id, "tank_type": data.tank_type, "liters": current_liters})
    return tank_dict

@api_router.get("/oxygen-tanks")
async def get_oxygen_tanks(
    ambulance_id: Optional[str] = None,
    is_portable: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    """Obtener tanques de oxígeno"""
    query = {}
    if ambulance_id:
        query["ambulance_id"] = ambulance_id
    if is_portable is not None:
        query["is_portable"] = is_portable
    
    tanks = await db.oxygen_tanks.find(query, {"_id": 0}).to_list(100)
    return tanks

@api_router.post("/oxygen-tanks/{tank_id}/movement")
async def create_oxygen_movement(
    tank_id: str,
    movement_type: OxygenMovementType,
    psi_after: int,
    observations: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Registrar movimiento de oxígeno"""
    tank = await db.oxygen_tanks.find_one({"id": tank_id})
    if not tank:
        raise HTTPException(status_code=404, detail="Tanque no encontrado")
    
    psi_before = tank["current_psi"]
    new_liters = calculate_liters_from_psi(psi_after, tank["max_psi"], tank["capacity_liters"])
    old_liters = tank["current_liters"]
    new_status = get_oxygen_tank_status(new_liters)
    
    # Calculate liters used/added
    liters_used = None
    liters_added = None
    if movement_type == OxygenMovementType.UTILIZADO:
        liters_used = old_liters - new_liters
    elif movement_type in [OxygenMovementType.RECARGADO, OxygenMovementType.REEMPLAZADO]:
        liters_added = new_liters - old_liters
    
    # Create movement record
    movement = {
        "id": str(uuid.uuid4()),
        "tank_id": tank_id,
        "ambulance_id": tank["ambulance_id"],
        "movement_type": movement_type,
        "psi_before": psi_before,
        "psi_after": psi_after,
        "liters_used": liters_used,
        "liters_added": liters_added,
        "observations": observations,
        "user_id": current_user.id,
        "user_name": current_user.full_name,
        "created_at": datetime.utcnow()
    }
    await db.oxygen_movements.insert_one(movement)
    
    # Update tank
    update_data = {
        "current_psi": psi_after,
        "current_liters": new_liters,
        "status": new_status,
        "last_movement_by": current_user.full_name,
        "updated_at": datetime.utcnow()
    }
    if movement_type in [OxygenMovementType.RECARGADO, OxygenMovementType.REEMPLAZADO]:
        update_data["last_refill_date"] = datetime.utcnow()
    
    await db.oxygen_tanks.update_one({"id": tank_id}, {"$set": update_data})
    
    # Get ambulance info for notifications
    ambulance = await db.ambulances.find_one({"id": tank["ambulance_id"]})
    unit = ambulance.get("unit_number", "N/A") if ambulance else "N/A"
    
    # Check for critical alerts
    if new_liters <= 0:
        # CRITICAL: Tank at 0
        await notify_all_coordinators(
            f"🚨🚨 EMERGENCIA: Oxígeno AGOTADO - Unidad {unit}",
            f"¡ATENCIÓN INMEDIATA! El tanque tipo {tank['tank_type']} está en 0 litros. Se requiere reemplazo URGENTE.",
            NotificationType.ALERTA_OXIGENO,
            NotificationPriority.CRITICA,
            entity_id=tank_id,
            entity_type="oxigeno"
        )
    elif new_liters <= 200 and old_liters > 200:
        # Status changed to CRITICAL
        await notify_all_coordinators(
            f"🚨 OXÍGENO CRÍTICO - Unidad {unit}",
            f"Tanque tipo {tank['tank_type']}: Solo {new_liters}L restantes. Recargar pronto.",
            NotificationType.ALERTA_OXIGENO,
            NotificationPriority.CRITICA,
            entity_id=tank_id,
            entity_type="oxigeno"
        )
    elif new_liters <= 800 and old_liters > 800:
        # Status changed to WARNING
        await notify_all_coordinators(
            f"⚠️ Oxígeno bajo - Unidad {unit}",
            f"Tanque tipo {tank['tank_type']}: {new_liters}L restantes ({psi_after} PSI)",
            NotificationType.ALERTA_OXIGENO,
            NotificationPriority.ALERTA,
            entity_id=tank_id,
            entity_type="oxigeno"
        )
    
    await log_audit(current_user.id, current_user.full_name, "movimiento", "oxigeno", tank_id,
                   {"movement_type": movement_type, "psi_before": psi_before, "psi_after": psi_after, 
                    "liters_before": old_liters, "liters_after": new_liters})
    
    return {
        "message": "Movimiento registrado",
        "tank": {**tank, **update_data},
        "movement": movement
    }

@api_router.get("/oxygen-tanks/{tank_id}/movements")
async def get_oxygen_movements(tank_id: str, current_user: User = Depends(get_current_user)):
    """Obtener historial de movimientos de un tanque"""
    movements = await db.oxygen_movements.find({"tank_id": tank_id}).sort("created_at", -1).to_list(100)
    return movements

# ==========================================
# INVENTORY ITEMS ROUTES
# ==========================================

@api_router.post("/inventory/initialize/{ambulance_id}")
async def initialize_ambulance_inventory(ambulance_id: str, current_user: User = Depends(get_current_user)):
    """Inicializar inventario completo para una ambulancia"""
    ambulance = await db.ambulances.find_one({"id": ambulance_id})
    if not ambulance:
        raise HTTPException(status_code=404, detail="Ambulancia no encontrada")
    
    # Check if already has inventory
    existing = await db.inventory.find_one({"ambulance_id": ambulance_id})
    if existing:
        raise HTTPException(status_code=400, detail="Esta ambulancia ya tiene inventario inicializado")
    
    items_created = []
    now = datetime.utcnow()
    
    for category, items in INVENTORY_TEMPLATE.items():
        for item in items:
            item_dict = {
                "id": str(uuid.uuid4()),
                "ambulance_id": ambulance_id,
                "name": item["name"],
                "category": category,
                "expected_quantity": item["expected"],
                "current_quantity": item["expected"],  # Start with full stock
                "unit": "piezas",
                "expiry_date": None,
                "lot_number": None,
                "observations": None,
                "has_difference": False,
                "is_below_minimum": False,
                "is_expiring_soon": False,
                "is_critical": item.get("is_critical", False),
                "requires_expiry": item.get("requires_expiry", False),
                "created_at": now,
                "updated_at": now
            }
            await db.inventory.insert_one(item_dict)
            items_created.append(item_dict)
    
    # Initialize oxygen tanks based on unit
    unit_number = ambulance.get("unit_number", "")
    
    if "UM05" in unit_number.upper() or "05" in unit_number:
        # UM05: 2 tanques tipo M (3500L)
        for i in range(2):
            tank = {
                "id": str(uuid.uuid4()),
                "ambulance_id": ambulance_id,
                "tank_type": OxygenTankType.M,
                "is_portable": False,
                "capacity_liters": 3500,
                "max_psi": 2000,
                "current_psi": 2000,
                "current_liters": 3500,
                "status": OxygenTankStatus.VERDE,
                "last_refill_date": now,
                "last_movement_by": current_user.full_name,
                "created_at": now,
                "updated_at": now
            }
            await db.oxygen_tanks.insert_one(tank)
    elif "UM03" in unit_number.upper() or "03" in unit_number:
        # UM03: 2 tanques tipo K (9500L)
        for i in range(2):
            tank = {
                "id": str(uuid.uuid4()),
                "ambulance_id": ambulance_id,
                "tank_type": OxygenTankType.K,
                "is_portable": False,
                "capacity_liters": 9500,
                "max_psi": 2000,
                "current_psi": 2000,
                "current_liters": 9500,
                "status": OxygenTankStatus.VERDE,
                "last_refill_date": now,
                "last_movement_by": current_user.full_name,
                "created_at": now,
                "updated_at": now
            }
            await db.oxygen_tanks.insert_one(tank)
    
    # Both units: 3 tanques tipo D + 1 tipo C (portátiles, 680L)
    for i in range(3):
        tank = {
            "id": str(uuid.uuid4()),
            "ambulance_id": ambulance_id,
            "tank_type": OxygenTankType.D,
            "is_portable": True,
            "capacity_liters": 680,
            "max_psi": 2000,
            "current_psi": 2000,
            "current_liters": 680,
            "status": OxygenTankStatus.VERDE,
            "last_refill_date": now,
            "last_movement_by": current_user.full_name,
            "created_at": now,
            "updated_at": now
        }
        await db.oxygen_tanks.insert_one(tank)
    
    # 1 tanque tipo C
    tank_c = {
        "id": str(uuid.uuid4()),
        "ambulance_id": ambulance_id,
        "tank_type": OxygenTankType.C,
        "is_portable": True,
        "capacity_liters": 680,
        "max_psi": 2000,
        "current_psi": 2000,
        "current_liters": 680,
        "status": OxygenTankStatus.VERDE,
        "last_refill_date": now,
        "last_movement_by": current_user.full_name,
        "created_at": now,
        "updated_at": now
    }
    await db.oxygen_tanks.insert_one(tank_c)
    
    await log_audit(current_user.id, current_user.full_name, "inicializar", "inventario", ambulance_id,
                   {"items_count": len(items_created)})
    
    return {"message": f"Inventario inicializado con {len(items_created)} items", "ambulance_id": ambulance_id}

@api_router.get("/inventory")
async def get_inventory(
    ambulance_id: Optional[str] = None,
    category: Optional[str] = None,
    with_difference: Optional[bool] = None,
    below_minimum: Optional[bool] = None,
    expiring_soon: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    """Obtener inventario con filtros"""
    query = {}
    if ambulance_id:
        query["ambulance_id"] = ambulance_id
    if category:
        query["category"] = category
    if with_difference:
        query["has_difference"] = True
    if below_minimum:
        query["is_below_minimum"] = True
    if expiring_soon:
        query["is_expiring_soon"] = True
    
    items = await db.inventory.find(query, {"_id": 0}).to_list(500)
    
    # Update expiry status for medications
    now = datetime.utcnow()
    thirty_days = timedelta(days=30)
    for item in items:
        if item.get("expiry_date"):
            expiry = item["expiry_date"]
            if isinstance(expiry, str):
                expiry = datetime.fromisoformat(expiry)
            item["is_expiring_soon"] = (expiry - now) <= thirty_days
            item["is_expired"] = expiry <= now
    
    return items

@api_router.get("/inventory/summary/{ambulance_id}")
async def get_inventory_summary(ambulance_id: str, current_user: User = Depends(get_current_user)):
    """Obtener resumen del inventario de una ambulancia"""
    items = await db.inventory.find({"ambulance_id": ambulance_id}).to_list(500)
    tanks = await db.oxygen_tanks.find({"ambulance_id": ambulance_id}).to_list(20)
    
    # Calculate stats
    total_items = len(items)
    items_with_difference = sum(1 for i in items if i.get("has_difference"))
    items_below_minimum = sum(1 for i in items if i.get("is_below_minimum"))
    
    # Check expiring items
    now = datetime.utcnow()
    thirty_days = timedelta(days=30)
    items_expiring_soon = 0
    for item in items:
        if item.get("expiry_date"):
            expiry = item["expiry_date"]
            if isinstance(expiry, str):
                expiry = datetime.fromisoformat(expiry)
            if (expiry - now) <= thirty_days:
                items_expiring_soon += 1
    
    # Oxygen summary
    portable_tanks = [t for t in tanks if t.get("is_portable")]
    stationary_tanks = [t for t in tanks if not t.get("is_portable")]
    
    portable_critical = sum(1 for t in portable_tanks if t.get("status") in ["rojo", "critico"])
    stationary_critical = sum(1 for t in stationary_tanks if t.get("status") in ["rojo", "critico"])
    
    return {
        "ambulance_id": ambulance_id,
        "total_items": total_items,
        "items_with_difference": items_with_difference,
        "items_below_minimum": items_below_minimum,
        "items_expiring_soon": items_expiring_soon,
        "oxygen_portable_count": len(portable_tanks),
        "oxygen_portable_critical": portable_critical,
        "oxygen_stationary_count": len(stationary_tanks),
        "oxygen_stationary_critical": stationary_critical,
        "categories": {
            cat.value: len([i for i in items if i.get("category") == cat.value])
            for cat in InventoryCategory
        }
    }

@api_router.post("/inventory/{item_id}/movement")
async def create_inventory_movement(
    item_id: str,
    movement_type: InventoryMovementType,
    quantity: int,
    reason: str,
    observations: Optional[str] = None,
    service_id: Optional[str] = None,
    is_correction: bool = False,
    correction_justification: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Registrar movimiento de inventario"""
    item = await db.inventory.find_one({"id": item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    
    quantity_before = item["current_quantity"]
    
    # Calculate new quantity
    if movement_type == InventoryMovementType.ENTRADA:
        quantity_after = quantity_before + quantity
    elif movement_type == InventoryMovementType.SALIDA:
        quantity_after = quantity_before - quantity
        if quantity_after < 0:
            raise HTTPException(status_code=400, detail="Stock insuficiente")
    elif movement_type in [InventoryMovementType.AJUSTE, InventoryMovementType.CORRECCION]:
        quantity_after = quantity  # Direct set
    else:
        quantity_after = quantity_before
    
    # Create movement record (cannot be deleted)
    movement = {
        "id": str(uuid.uuid4()),
        "item_id": item_id,
        "item_name": item["name"],
        "ambulance_id": item["ambulance_id"],
        "movement_type": movement_type,
        "quantity_before": quantity_before,
        "quantity_moved": quantity if movement_type in [InventoryMovementType.ENTRADA, InventoryMovementType.SALIDA] else abs(quantity_after - quantity_before),
        "quantity_after": quantity_after,
        "reason": reason,
        "observations": observations,
        "service_id": service_id,
        "user_id": current_user.id,
        "user_name": current_user.full_name,
        "created_at": datetime.utcnow(),
        "is_correction": is_correction,
        "correction_justification": correction_justification
    }
    await db.inventory_movements.insert_one(movement)
    
    # Update item
    has_difference = quantity_after != item["expected_quantity"]
    is_below_minimum = quantity_after < (item["expected_quantity"] * 0.3)  # Below 30% is minimum
    
    await db.inventory.update_one(
        {"id": item_id},
        {"$set": {
            "current_quantity": quantity_after,
            "has_difference": has_difference,
            "is_below_minimum": is_below_minimum,
            "updated_at": datetime.utcnow()
        }}
    )
    
    # Get ambulance info
    ambulance = await db.ambulances.find_one({"id": item["ambulance_id"]})
    unit = ambulance.get("unit_number", "N/A") if ambulance else "N/A"
    
    # Check for alerts
    if is_below_minimum and not item.get("is_below_minimum"):
        is_critical = item.get("is_critical", False)
        priority = NotificationPriority.CRITICA if is_critical else NotificationPriority.ALERTA
        await notify_all_coordinators(
            f"{'🚨 CRÍTICO' if is_critical else '⚠️'}: Stock bajo - {item['name']}",
            f"Unidad {unit}: {item['name']} en {quantity_after}/{item['expected_quantity']}",
            NotificationType.ALERTA_INVENTARIO,
            priority,
            entity_id=item_id,
            entity_type="inventario"
        )
    
    # Log medication movements
    if item.get("category") == InventoryCategory.MEDICAMENTOS and movement_type == InventoryMovementType.SALIDA:
        await notify_all_coordinators(
            f"💊 Salida de medicamento - Unidad {unit}",
            f"{item['name']}: {quantity} unidades utilizadas. Stock actual: {quantity_after}",
            NotificationType.ALERTA_INVENTARIO,
            NotificationPriority.NORMAL,
            entity_id=item_id,
            entity_type="inventario"
        )
    
    await log_audit(current_user.id, current_user.full_name, "movimiento", "inventario", item_id,
                   {"movement_type": movement_type, "quantity_before": quantity_before, 
                    "quantity_after": quantity_after, "reason": reason})
    
    return {
        "message": "Movimiento registrado",
        "movement": movement,
        "item": {**item, "current_quantity": quantity_after, "has_difference": has_difference, "is_below_minimum": is_below_minimum}
    }

@api_router.put("/inventory/{item_id}")
async def update_inventory_item(
    item_id: str,
    current_quantity: Optional[int] = None,
    expiry_date: Optional[str] = None,
    lot_number: Optional[str] = None,
    observations: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Actualizar item de inventario"""
    item = await db.inventory.find_one({"id": item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    
    update_data = {"updated_at": datetime.utcnow()}
    
    if current_quantity is not None:
        update_data["current_quantity"] = current_quantity
        update_data["has_difference"] = current_quantity != item["expected_quantity"]
        update_data["is_below_minimum"] = current_quantity < (item["expected_quantity"] * 0.3)
    
    if expiry_date:
        update_data["expiry_date"] = datetime.fromisoformat(expiry_date)
        # Check if expiring soon
        thirty_days = timedelta(days=30)
        update_data["is_expiring_soon"] = (datetime.fromisoformat(expiry_date) - datetime.utcnow()) <= thirty_days
    
    if lot_number:
        update_data["lot_number"] = lot_number
    
    if observations is not None:
        update_data["observations"] = observations
    
    await db.inventory.update_one({"id": item_id}, {"$set": update_data})
    
    await log_audit(current_user.id, current_user.full_name, "actualizar", "inventario", item_id, update_data)
    
    updated_item = await db.inventory.find_one({"id": item_id})
    return updated_item

@api_router.get("/inventory/{item_id}/movements")
async def get_item_movements(
    item_id: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Obtener historial de movimientos de un item"""
    query = {"item_id": item_id}
    
    if date_from:
        query["created_at"] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(date_to)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(date_to)}
    
    movements = await db.inventory_movements.find(query).sort("created_at", -1).to_list(200)
    return movements

@api_router.get("/inventory/movements/ambulance/{ambulance_id}")
async def get_ambulance_movements(
    ambulance_id: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    paramedic_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Obtener todos los movimientos de inventario de una ambulancia"""
    query = {"ambulance_id": ambulance_id}
    
    if date_from:
        query["created_at"] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(date_to)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(date_to)}
    if paramedic_id:
        query["user_id"] = paramedic_id
    
    movements = await db.inventory_movements.find(query).sort("created_at", -1).to_list(500)
    return movements

# ==========================================
# INVENTORY RECORD (CHECKLIST STYLE)
# ==========================================

@api_router.post("/inventory/record")
async def create_inventory_record(
    ambulance_id: str,
    shift: str,
    paramedic_id: str,
    observations: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Crear registro de revisión de inventario (tipo checklist)"""
    ambulance = await db.ambulances.find_one({"id": ambulance_id})
    if not ambulance:
        raise HTTPException(status_code=404, detail="Ambulancia no encontrada")
    
    paramedic = await db.personnel.find_one({"id": paramedic_id})
    if not paramedic:
        raise HTTPException(status_code=404, detail="Paramédico no encontrado")
    
    # Get current inventory state
    items = await db.inventory.find({"ambulance_id": ambulance_id}).to_list(500)
    tanks = await db.oxygen_tanks.find({"ambulance_id": ambulance_id}).to_list(20)
    
    # Calculate stats
    items_with_diff = sum(1 for i in items if i.get("has_difference"))
    items_below_min = sum(1 for i in items if i.get("is_below_minimum"))
    items_expiring = sum(1 for i in items if i.get("is_expiring_soon"))
    
    # Oxygen status summary
    portable_tanks = [t for t in tanks if t.get("is_portable")]
    stationary_tanks = [t for t in tanks if not t.get("is_portable")]
    
    portable_status = "🟢 OK"
    stationary_status = "🟢 OK"
    
    for t in portable_tanks:
        if t.get("status") in ["critico", "rojo"]:
            portable_status = "🔴 CRÍTICO"
            break
        elif t.get("status") == "amarillo":
            portable_status = "🟡 BAJO"
    
    for t in stationary_tanks:
        if t.get("status") in ["critico", "rojo"]:
            stationary_status = "🔴 CRÍTICO"
            break
        elif t.get("status") == "amarillo":
            stationary_status = "🟡 BAJO"
    
    record = {
        "id": str(uuid.uuid4()),
        "ambulance_id": ambulance_id,
        "ambulance_unit": ambulance.get("unit_number", "N/A"),
        "date": datetime.utcnow(),
        "shift": shift,
        "paramedic_id": paramedic_id,
        "paramedic_name": paramedic.get("full_name", "N/A"),
        "oxygen_portable_status": portable_status,
        "oxygen_stationary_status": stationary_status,
        "total_items": len(items),
        "items_with_difference": items_with_diff,
        "items_below_minimum": items_below_min,
        "items_expiring_soon": items_expiring,
        "observations": observations,
        "created_at": datetime.utcnow(),
        "created_by": current_user.id
    }
    
    await db.inventory_records.insert_one(record)
    
    # Generate alerts if needed
    if items_with_diff > 0 or items_below_min > 0:
        await notify_all_coordinators(
            f"⚠️ Revisión inventario - Unidad {ambulance.get('unit_number', 'N/A')}",
            f"Turno {shift}: {items_with_diff} items con diferencia, {items_below_min} bajo mínimo",
            NotificationType.ALERTA_INVENTARIO,
            NotificationPriority.ALERTA if items_below_min > 0 else NotificationPriority.NORMAL,
            entity_id=record["id"],
            entity_type="inventario_record"
        )
    
    await log_audit(current_user.id, current_user.full_name, "crear", "registro_inventario", record["id"],
                   {"ambulance_id": ambulance_id, "shift": shift, "items_diff": items_with_diff})
    
    return record

@api_router.get("/inventory/records")
async def get_inventory_records(
    ambulance_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Obtener registros de revisión de inventario"""
    query = {}
    if ambulance_id:
        query["ambulance_id"] = ambulance_id
    if date_from:
        query["date"] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        if "date" in query:
            query["date"]["$lte"] = datetime.fromisoformat(date_to)
        else:
            query["date"] = {"$lte": datetime.fromisoformat(date_to)}
    
    records = await db.inventory_records.find(query).sort("date", -1).to_list(100)
    return records

# Audit Logs
@api_router.get("/audit-logs")
async def get_audit_logs(
    entity_type: Optional[str] = None,
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Only supervisors and admins can view audit logs
    if current_user.role not in ["supervisor", "coordinador", "administrativo"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if user_id:
        query["user_id"] = user_id
    if date_from:
        query["created_at"] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(date_to)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(date_to)}
    
    logs = await db.audit_logs.find(query).sort("created_at", -1).to_list(500)
    return logs

# PDF Report Generation
@api_router.get("/reports/weekly-pdf")
async def generate_weekly_report(
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user)
):
    query = {
        "created_at": {
            "$gte": datetime.fromisoformat(date_from),
            "$lte": datetime.fromisoformat(date_to)
        }
    }
    entries = await db.finances.find(query).to_list(1000)
    
    service_query = {
        "scheduled_date": {
            "$gte": datetime.fromisoformat(date_from),
            "$lte": datetime.fromisoformat(date_to)
        }
    }
    services = await db.services.find(service_query).to_list(1000)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,
        textColor=colors.HexColor('#1E40AF')
    )
    elements.append(Paragraph("AMBULANCIAS STEN", title_style))
    elements.append(Paragraph("CORTE SEMANAL", styles['Heading2']))
    elements.append(Paragraph(f"Período: {date_from} al {date_to}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    total_ingresos = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso")
    total_egresos = sum(e["amount"] for e in entries if e["entry_type"] == "egreso")
    ingresos_efectivo = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "efectivo")
    ingresos_transferencia = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "transferencia")
    ingresos_tarjeta = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "tarjeta")
    ingresos_hospital = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "hospital")
    balance = total_ingresos - total_egresos
    
    summary_data = [
        ["RESUMEN FINANCIERO", ""],
        ["Total Ingresos:", f"${total_ingresos:,.2f}"],
        ["  - Efectivo:", f"${ingresos_efectivo:,.2f}"],
        ["  - Transferencia:", f"${ingresos_transferencia:,.2f}"],
        ["  - Tarjeta (+8%):", f"${ingresos_tarjeta:,.2f}"],
        ["  - Hospital:", f"${ingresos_hospital:,.2f}"],
        ["Total Egresos:", f"${total_egresos:,.2f}"],
        ["BALANCE:", f"${balance:,.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E5E7EB'))
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Services summary
    total_services = len(services)
    completed = len([s for s in services if s.get("status") == "finalizado"])
    urgent = len([s for s in services if s.get("service_type") == "urgente"])
    elements.append(Paragraph(f"Total de servicios: {total_services} | Completados: {completed} | Urgentes: {urgent}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Income details
    if any(e["entry_type"] == "ingreso" for e in entries):
        elements.append(Paragraph("DETALLE DE INGRESOS", styles['Heading2']))
        income_data = [["Fecha", "Descripción", "Tipo Pago", "Monto"]]
        for e in entries:
            if e["entry_type"] == "ingreso":
                income_data.append([
                    e["created_at"].strftime("%Y-%m-%d"),
                    e["description"][:30],
                    e.get("payment_type", "N/A"),
                    f"${e['amount']:,.2f}"
                ])
        
        income_table = Table(income_data, colWidths=[1.2*inch, 2.5*inch, 1.3*inch, 1*inch])
        income_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16A34A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        elements.append(income_table)
        elements.append(Spacer(1, 20))
    
    # Expense details
    if any(e["entry_type"] == "egreso" for e in entries):
        elements.append(Paragraph("DETALLE DE EGRESOS", styles['Heading2']))
        expense_data = [["Fecha", "Descripción", "Categoría", "Monto"]]
        for e in entries:
            if e["entry_type"] == "egreso":
                expense_data.append([
                    e["created_at"].strftime("%Y-%m-%d"),
                    e["description"][:30],
                    e.get("category", "N/A"),
                    f"${e['amount']:,.2f}"
                ])
        
        expense_table = Table(expense_data, colWidths=[1.2*inch, 2.5*inch, 1.3*inch, 1*inch])
        expense_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        elements.append(expense_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=corte_semanal_{date_from}_{date_to}.pdf"}
    )

# Excel Report Generation
@api_router.get("/reports/weekly-excel")
async def generate_weekly_excel(
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user)
):
    query = {
        "created_at": {
            "$gte": datetime.fromisoformat(date_from),
            "$lte": datetime.fromisoformat(date_to)
        }
    }
    entries = await db.finances.find(query).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Corte Semanal"
    
    # Styles
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    income_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    expense_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = f"AMBULANCIAS STEN - Corte Semanal ({date_from} al {date_to})"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Summary
    total_ingresos = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso")
    total_egresos = sum(e["amount"] for e in entries if e["entry_type"] == "egreso")
    
 ingresos_efectivo = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "efectivo")
    ingresos_transferencia = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "transferencia")
    ingresos_tarjeta = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "tarjeta")
    ingresos_hospital = sum(e["amount"] for e in entries if e["entry_type"] == "ingreso" and e.get("payment_type") == "hospital")

    ws['A3'] = "Total Ingresos:"
    ws['B3'] = total_ingresos
    ws['B3'].number_format = '$#,##0.00'
    ws['A4'] = "  - Efectivo:"
    ws['B4'] = ingresos_efectivo
    ws['B4'].number_format = '$#,##0.00'
    ws['A5'] = "  - Transferencia:"
    ws['B5'] = ingresos_transferencia
    ws['B5'].number_format = '$#,##0.00'
    ws['A6'] = "  - Tarjeta (+8%):"
    ws['B6'] = ingresos_tarjeta
    ws['B6'].number_format = '$#,##0.00'
    ws['A7'] = "  - Hospital:"
    ws['B7'] = ingresos_hospital
    ws['B7'].number_format = '$#,##0.00'
    ws['A8'] = "Total Egresos:"
    ws['B8'] = total_egresos
    ws['B8'].number_format = '$#,##0.00'
    ws['A9'] = "Balance:"
    ws['B9'] = total_ingresos - total_egresos
    ws['B9'].number_format = '$#,##0.00'
    ws['B9'].font = Font(bold=True)
    
    # Headers
    headers = ["Fecha", "Tipo", "Descripción", "Categoría", "Monto"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    row = 8
    for entry in entries:
        ws.cell(row=row, column=1, value=entry["created_at"].strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value=entry["entry_type"].upper())
        ws.cell(row=row, column=3, value=entry["description"])
        ws.cell(row=row, column=4, value=entry.get("category", entry.get("payment_type", "")))
        amount_cell = ws.cell(row=row, column=5, value=entry["amount"])
        amount_cell.number_format = '$#,##0.00'
        
        if entry["entry_type"] == "ingreso":
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        else:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=corte_semanal_{date_from}_{date_to}.xlsx"}
    )

# Dashboard Stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    
    total_services = await db.services.count_documents({})
    pending_services = await db.services.count_documents({"status": "pendiente"})
    confirmed_services = await db.services.count_documents({"status": "confirmado"})
    en_camino_services = await db.services.count_documents({"status": "en_camino"})
    in_progress = await db.services.count_documents({"status": "en_curso"})
    today_services = await db.services.count_documents({
        "scheduled_date": {"$gte": today, "$lt": today + timedelta(days=1)}
    })
    urgent_pending = await db.services.count_documents({
        "service_type": "urgente",
        "status": {"$in": ["pendiente", "confirmado"]}
    })
    
    total_ambulances = await db.ambulances.count_documents({})
    available_ambulances = await db.ambulances.count_documents({"status": "disponible"})
    in_service_ambulances = await db.ambulances.count_documents({"status": "en_servicio"})
    maintenance_ambulances = await db.ambulances.count_documents({"status": {"$in": ["mantenimiento", "fuera_servicio"]}})
    
    total_personnel = await db.personnel.count_documents({})
    available_personnel = await db.personnel.count_documents({"is_available": True})
    
    unread_notifications = await db.notifications.count_documents({
        "user_id": current_user.id,
        "is_read": False
    })
    
    # Low stock items
    low_stock_items = await db.inventory.find().to_list(500)
    low_stock_count = sum(1 for i in low_stock_items if i.get("current_stock", 0) <= i.get("min_stock", 0))
    
    return {
        "services": {
            "total": total_services,
            "pending": pending_services,
            "confirmed": confirmed_services,
            "en_camino": en_camino_services,
            "in_progress": in_progress,
            "today": today_services,
            "urgent_pending": urgent_pending
        },
        "ambulances": {
            "total": total_ambulances,
            "available": available_ambulances,
            "in_service": in_service_ambulances,
            "maintenance": maintenance_ambulances
        },
        "personnel": {
            "total": total_personnel,
            "available": available_personnel
        },
        "unread_notifications": unread_notifications,
        "low_stock_items": low_stock_count
    }

# Public Emergency Endpoint
@api_router.post("/public/emergency")
async def create_emergency_request(data: ServiceCreate):
    service_dict = {
        "id": str(uuid.uuid4()),
        **data.dict(),
        "service_type": ServiceType.URGENTE,
        "status": ServiceStatus.PENDIENTE,
        "ambulance_id": None,
        "personnel_ids": [],
        "payment_type": PaymentType.PENDIENTE,
        "payment_amount": None,
        "frap_notes": None,
        "bank_name": None,
        "account_number": None,
        "account_holder": None,
        "transfer_reference": None,
        "payment_proof_base64": None,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow(),
        "created_by": "public_qr",
        "status_history": [{
            "status": "pendiente",
            "timestamp": datetime.utcnow().isoformat(),
            "user_id": "public",
            "user_name": "Solicitud Pública"
        }]
    }
    await db.services.insert_one(service_dict)
    
    await notify_all_coordinators(
        "🚨 URGENCIA - Nueva solicitud pública",
        f"Paciente: {data.patient.name} - Origen: {data.origin}",
        NotificationType.URGENCIA,
        service_dict["id"]
    )
    
    await check_available_ambulances()
    
    return {"message": "Solicitud de emergencia recibida", "service_id": service_dict["id"]}

# Health check
@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "app": "Ambulancias STEN", "timestamp": datetime.utcnow().isoformat()}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
